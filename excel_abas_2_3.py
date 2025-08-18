#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gerador Excel - Abas 2 e 3 do Sistema FIDC Energisa
====================================================
Aba 2: Dicionário de Dados
Aba 3: Cálculos e Metodologia
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import os

def criar_excel_documentacao():
    """
    Cria arquivo Excel com documentação completa
    """
    print("🔋 ENERGISA FIDC - Gerando Excel Documentação")
    print("=" * 50)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Remover aba padrão
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Estilos
    estilo_cabecalho = {
        'font': Font(bold=True, color="FFFFFF", size=12),
        'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    estilo_dados = {
        'font': Font(size=10),
        'alignment': Alignment(horizontal="left", vertical="center"),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    # ============ ABA 2: DICIONÁRIO DE DADOS ============
    ws2 = wb.create_sheet("2. Dicionário de Dados")
    
    # Título
    ws2['A1'] = "📚 DICIONÁRIO DE DADOS - DataFrame Final FIDC"
    ws2['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws2.merge_cells('A1:E1')
    
    # Cabeçalhos
    headers = ['#', 'Nome da Coluna', 'Tipo de Dado', 'Descrição', 'Observações']
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = estilo_cabecalho['font']
        cell.fill = estilo_cabecalho['fill']
        cell.alignment = estilo_cabecalho['alignment']
        cell.border = estilo_cabecalho['border']
    
    # Dicionário de dados
    dicionario_dados = [
        # GRUPO 1: Dados Básicos do Cliente
        (1, "nome_cliente", "Texto", "Nome do cliente titular da conta", "Campo obrigatório para identificação"),
        (2, "documento", "Numérico", "CPF/CNPJ do cliente", "Frequentemente vazio nos dados"),
        (3, "contrato", "Numérico", "Número do contrato de energia", "Identificador único do contrato"),
        (4, "classe", "Texto", "Classificação do cliente", "Residencial, Comercial, Industrial, Rural"),
        (5, "situacao", "Texto", "Status atual do cliente", "Ativo, Desligado, Suspenso"),
        (6, "valor_principal", "Numérico", "Valor original da dívida", "Valor base antes de deduções"),
        (7, "valor_nao_cedido", "Numérico", "Parcela não cedida ao FIDC", "Valor retido pela distribuidora"),
        (8, "valor_terceiro", "Numérico", "Valores de terceiros", "Tributos, taxas de terceiros"),
        (9, "valor_cip", "Numérico", "Valor CIP", "Contribuição de Iluminação Pública"),
        
        # GRUPO 2: Metadados e Controle
        (10, "data_vencimento", "Data", "Data original de vencimento", "Formato timestamp original"),
        (11, "empresa", "Texto", "Código da distribuidora", "ESS, EMR, etc."),
        (12, "tipo", "Texto", "Classificação do negócio", "Privado, Público, Hospital"),
        (13, "status", "Texto", "Status da cobrança", "Cobrável, Incobrável, Em análise"),
        (14, "id_padronizado", "Texto", "Identificador único gerado", "Chave primária para rastreamento"),
        (15, "base_origem", "Texto", "Arquivo de origem dos dados", "Rastreabilidade da fonte"),
        (16, "data_base", "Data", "Data de referência para cálculos", "Data fixa para todos os cálculos"),
        
        # GRUPO 3: Aging e Temporalidade
        (17, "data_vencimento_limpa", "Data", "Data de vencimento padronizada", "Formato YYYY-MM-DD limpo"),
        (18, "dias_atraso", "Inteiro", "Dias entre vencimento e data base", "Cálculo: data_base - data_vencimento"),
        (19, "aging", "Texto", "Classificação de aging", "9 faixas: A vencer até >1080 dias"),
        
        # GRUPO 4: Valores Limpos e Líquidos
        (20, "valor_principal_limpo", "Numérico", "Valor principal sem formatação", "Numérico puro"),
        (21, "valor_nao_cedido_limpo", "Numérico", "Valor não cedido limpo", "Numérico puro"),
        (22, "valor_terceiro_limpo", "Numérico", "Valor terceiros limpo", "Numérico puro"),
        (23, "valor_cip_limpo", "Numérico", "Valor CIP limpo", "Numérico puro"),
        (24, "valor_liquido", "Numérico", "Valor líquido final", "Fórmula: principal - nao_cedido - terceiro - cip"),
        
        # GRUPO 5: Correção Monetária
        (25, "multa", "Numérico", "Multa calculada", "Fórmula: valor_liquido × 2%"),
        (26, "meses_atraso", "Numérico", "Meses de atraso calculados", "dias_atraso ÷ 30"),
        (27, "juros_moratorios", "Numérico", "Juros moratórios", "Fórmula: valor_liquido × 1% × meses"),
        (28, "indice_vencimento", "Numérico", "Índice IGP-M no vencimento", "Obtido da tabela histórica"),
        (29, "indice_base", "Numérico", "Índice IGP-M na data base", "Índice de referência (data_base)"),
        (30, "fator_correcao", "Numérico", "Fator de correção monetária", "Fórmula: indice_base ÷ indice_vencimento"),
        (31, "correcao_monetaria", "Numérico", "Correção monetária aplicada", "Fórmula: valor_liquido × (fator_correcao - 1)"),
        (32, "valor_corrigido", "Numérico", "Valor final corrigido", "Fórmula: valor_liquido + multa + juros + correção"),
        
        # GRUPO 6: Taxa de Recuperação e Valor Justo
        (33, "aging_taxa", "Texto", "Aging mapeado para taxa", "Mapeamento: >1080 dias → Demais anos"),
        (34, "taxa_recuperacao", "Numérico", "Taxa de recuperação esperada", "Taxa decimal obtida via merge por Empresa×Tipo×Aging"),
        (35, "prazo_recebimento", "Numérico", "Prazo esperado de recebimento", "Meses até recebimento (da tabela recuperação)"),
        (36, "valor_recuperavel", "Numérico", "Valor esperado de recuperação", "Fórmula: valor_corrigido × taxa_recuperacao"),
        
        # GRUPO 7: Cálculo DI-PRE e Valor Justo
        (37, "taxa_di_pre", "Numérico", "Taxa DI-PRE da BMF", "Taxa extraída do arquivo BMF para prazo específico"),
        (38, "fator_exponencial_di_pre", "Numérico", "Fator exponencial DI-PRE", "Fórmula: (1 + taxa_di_pre)^(prazo_recebimento/12)"),
        (39, "data_vencimento", "Data", "Data vencimento para valor justo", "Data base + 6 meses (referência para multa)"),
        (40, "dias_atraso", "Inteiro", "Dias de atraso para valor justo", "Diferença entre hoje e data_vencimento"),
        (41, "multa_para_justo", "Numérico", "Multa proporcional valor justo", "0.01/30 × dias_atraso ou 6% se atraso = 0"),
        (42, "valor_justo", "Numérico", "Valor justo final com DI-PRE", "VC × Taxa_Rec × (Fator_Exp_DI_PRE + Multa_Prop)")
    ]
    
    # Inserir dados do dicionário
    for row_idx, (num, coluna, tipo, descricao, obs) in enumerate(dicionario_dados, 4):
        # Aplicar cores alternadas
        fill_color = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
        
        cells = [
            ws2.cell(row=row_idx, column=1, value=num),
            ws2.cell(row=row_idx, column=2, value=coluna),
            ws2.cell(row=row_idx, column=3, value=tipo),
            ws2.cell(row=row_idx, column=4, value=descricao),
            ws2.cell(row=row_idx, column=5, value=obs)
        ]
        
        for cell in cells:
            cell.border = estilo_dados['border']
            cell.alignment = estilo_dados['alignment']
            cell.font = estilo_dados['font']
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    
    # Ajustar largura das colunas
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 40
    ws2.column_dimensions['E'].width = 35
    
    # ============ ABA 3: CÁLCULOS E METODOLOGIA ============
    ws3 = wb.create_sheet("3. Cálculos e Metodologia")
    
    # Título
    ws3['A1'] = "🧮 METODOLOGIA DE CÁLCULOS FIDC - ENERGISA"
    ws3['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws3.merge_cells('A1:F1')
    
    linha_atual = 3
    
    # Seções explicativas
    secoes = [
        {
            'titulo': '1. VISÃO GERAL DO SISTEMA',
            'conteudo': [
                'O sistema FIDC Calculator implementa metodologia rigorosa para cálculo de valor justo',
                'de carteiras de recebíveis de distribuidoras de energia elétrica brasileiras.',
                '',
                'COMPONENTES PRINCIPAIS:',
                '• Aging (classificação temporal da inadimplência)',
                '• Correção monetária (IGP-M até mai/2021, IPCA após jun/2021)',
                '• Juros moratórios (1% ao mês, proporcional)',
                '• Multa contratual (2% sobre valor líquido)',
                '• Taxa de recuperação (baseada em histórico setorial)',
                '• Valor presente (desconto pelo prazo de recebimento)'
            ]
        },
        {
            'titulo': '2. CÁLCULO DE AGING',
            'conteudo': [
                'OBJETIVO: Classificar cada dívida pelo tempo de inadimplência',
                '',
                'FÓRMULA: Dias de Atraso = Data Base - Data de Vencimento',
                '',
                'CLASSIFICAÇÕES:',
                '• A vencer: DA ≤ 0 (ainda não vencido)',
                '• Menor que 30 dias: 0 < DA ≤ 30',
                '• De 31 a 59 dias: 30 < DA ≤ 59',
                '• De 60 a 89 dias: 59 < DA ≤ 89',
                '• De 90 a 119 dias: 89 < DA ≤ 119',
                '• De 120 a 359 dias: 119 < DA ≤ 359 (CONCENTRAÇÃO DE VALOR)',
                '• De 360 a 719 dias: 359 < DA ≤ 719',
                '• De 720 a 1080 dias: 719 < DA ≤ 1080',
                '• Maior que 1080 dias: DA > 1080'
            ]
        },
        {
            'titulo': '3. CORREÇÃO MONETÁRIA',
            'conteudo': [
                'OBJETIVO: Recompor poder de compra considerando inflação oficial',
                '',
                'METODOLOGIA HÍBRIDA ATUAL:',
                '• IGP-M: para datas ≤ maio/2021 (índice base 624.40)',
                '• IPCA: para datas > maio/2021 (índice base 1069.29)',
                '',
                'FÓRMULA: CM = Valor Líquido × (Índice Base ÷ Índice Vencimento - 1)',
                '',
                'ÍNDICES UTILIZADOS:',
                '• IGP-M: Base histórica até maio/2021',
                '• IPCA: Via API SIDRA/IBGE a partir de junho/2021',
                '',
                'APLICAÇÃO: Apenas para valores em atraso (dias_atraso > 0)',
                'PROTEÇÃO: Correção sempre ≥ 0 (sem deflação)'
            ]
        },
        {
            'titulo': '4. JUROS E MULTAS',
            'conteudo': [
                'MULTA CONTRATUAL:',
                '• Taxa: 2% sobre valor líquido',
                '• Aplicação: Uma única vez para valores em atraso',
                '• Fórmula: M = VL × 0,02 × I(DA > 0)',
                '',
                'JUROS MORATÓRIOS:',
                '• Taxa: 1% ao mês',
                '• Cálculo proporcional por dias',
                '• Fórmula: J = VL × 0,01 × (DA ÷ 30) × I(DA > 0)',
                '',
                'VALOR CORRIGIDO TOTAL:',
                'VC = Valor Líquido + Multa + Juros + Correção Monetária'
            ]
        },
        {
            'titulo': '5. TAXAS DE RECUPERAÇÃO',
            'conteudo': [
                'OBJETIVO: Estimar probabilidade de recuperação efetiva por aging',
                '',
                'MAPEAMENTO DE AGING PARA TAXAS:',
                '• A vencer → A vencer',
                '• Menor que 30 dias → Primeiro ano', 
                '• De 31 a 59 dias → Primeiro ano',
                '• De 60 a 89 dias → Primeiro ano',
                '• De 90 a 119 dias → Primeiro ano',
                '• De 120 a 359 dias → Primeiro ano (CONCENTRAÇÃO)',
                '• De 360 a 719 dias → Segundo ano',
                '• De 720 a 1080 dias → Terceiro ano',
                '• Maior que 1080 dias → Demais anos',
                '',
                'FONTE: Tabela de recuperação por Empresa × Tipo × Aging',
                'APLICAÇÃO: Merge automático baseado em chaves compostas',
                'FALLBACK: Taxa 0% para combinações não encontradas'
            ]
        },
        {
            'titulo': '6. VALOR JUSTO (CÁLCULO COM DI-PRE)',
            'conteudo': [
                'OBJETIVO: Determinar valor presente usando taxas DI-PRE da BMF',
                '',
                'METODOLOGIA ATUALIZADA:',
                '• Utiliza dados reais de DI×PRE da BMF Bovespa',
                '• Aplicação de progressão exponencial por prazo',
                '• Incorporação de multa proporcional por atraso',
                '',
                'FÓRMULA COMPLETA:',
                'VJ = Valor_Corrigido × Taxa_Recuperação × (Fator_Exponencial_DI_PRE + Multa_Proporcional)',
                '',
                'COMPONENTES:',
                '• Fator_Exponencial_DI_PRE = (1 + taxa_di_pre)^(prazo_recebimento/12)',
                '• Multa_Proporcional = 0.01/30 × dias_atraso (fallback: 6% se atraso = 0)',
                '• Taxa_DI_PRE: extraída de arquivo BMF para prazo específico',
                '',
                'PRAZOS DE RECEBIMENTO:',
                '• Baseados na tabela de recuperação por aging',
                '• Variação típica: 6 a 42 meses',
                '',
                'FONTE: Arquivos PRE (DI×PRE) da BMF Bovespa'
            ]
        },
        {
            'titulo': '7. ENCADEAMENTO DOS CÁLCULOS ATUAL',
            'conteudo': [
                'SEQUÊNCIA LÓGICA IMPLEMENTADA:',
                '',
                '1. VALOR LÍQUIDO = Principal - Não Cedido - Terceiros - CIP',
                '2. AGING = Classificação por dias de atraso (9 faixas)',
                '3. MULTA = VL × 2% (apenas se dias_atraso > 0)',
                '4. JUROS = VL × 1% × (dias_atraso ÷ 30) (apenas se atraso > 0)',
                '5. CORREÇÃO = VL × (Índice_Base ÷ Índice_Vencimento - 1)',
                '6. VALOR CORRIGIDO = VL + Multa + Juros + Correção',
                '7. AGING_TAXA = Mapeamento aging → categoria recuperação',
                '8. TAXA_RECUPERAÇÃO = Merge com tabela (Empresa × Tipo × Aging)',
                '9. VALOR_RECUPERÁVEL = VC × Taxa_Recuperação',
                '10. TAXA_DI_PRE = Busca em arquivo BMF por prazo_recebimento',
                '11. FATOR_EXPONENCIAL = (1 + taxa_di_pre)^(prazo/12)',
                '12. MULTA_PROPORCIONAL = (0.01/30) × dias_atraso ou 6%',
                '13. VALOR_JUSTO = VC × Taxa_Rec × (Fator_Exp + Multa_Prop)',
                '',
                'RESULTADO: Valor justo considerando cenário macroeconômico atual'
            ]
        },
        {
            'titulo': '8. FATORES DE INFLUÊNCIA E INOVAÇÕES',
            'conteudo': [
                'COMPOSIÇÃO DA CARTEIRA:',
                '• Concentração em aging 120-359 dias ainda maximiza valor',
                '• Carteiras antigas (>3 anos) têm recuperação limitada',
                '• Mix equilibrado otimiza relação risco-retorno',
                '',
                'INOVAÇÕES DO SISTEMA ATUAL:',
                '• Uso de taxas DI-PRE reais da BMF em tempo real',
                '• Cálculo exponencial baseado em prazo específico',
                '• Multa proporcional por dias de atraso efetivos',
                '• Mapeamento automático aging → categoria recuperação',
                '',
                'CONDIÇÕES MACROECONÔMICAS:',
                '• Taxa DI-PRE reflete cenário atual do mercado',
                '• IPCA atualizado via API SIDRA para máxima precisão',
                '• Progressão exponencial captura efeito compound',
                '',
                'CARACTERÍSTICAS SETORIAIS:',
                '• Essencialidade do serviço mantém taxas de recuperação',
                '• Regulamentação específica influencia prazos de cobrança',
                '• Sazonalidade considerada nos prazos de recebimento'
            ]
        },
        {
            'titulo': '9. RESULTADOS ESPERADOS E VALIDAÇÃO',
            'conteudo': [
                'FAIXAS DE VALOR JUSTO ATUALIZADAS:',
                '',
                'POR AGING (com DI-PRE atual):',
                '• Primeiro ano: Taxa DI-PRE × progressão exponencial',
                '• Segundo ano: Taxas menores devido ao maior prazo',
                '• Terceiro ano: Impacto significativo da progressão temporal',
                '• Acima de 3 anos: Valores residuais com alta incerteza',
                '',
                'VALIDAÇÃO DO SISTEMA:',
                '• Comparação com benchmarks de mercado FIDC',
                '• Análise de sensibilidade por cenário macroeconômico',
                '• Validação cruzada Empresa × Tipo × Aging',
                '',
                'OUTPUTS DE CONTROLE:',
                '• Percentual de registros com taxa de recuperação aplicada',
                '• Estatísticas de match automático aging → categoria',
                '• Distribuição de prazos de recebimento por carteira',
                '',
                'MÉTRICAS DE QUALIDADE:',
                'Sistema reporta % de sucesso em cada etapa de cálculo'
            ]
        }
    ]
    
    # Inserir seções
    for secao in secoes:
        # Título da seção
        ws3.cell(row=linha_atual, column=1, value=secao['titulo']).font = Font(bold=True, size=12, color="1F4E79")
        linha_atual += 2
        
        # Conteúdo da seção
        for linha_conteudo in secao['conteudo']:
            ws3.cell(row=linha_atual, column=1, value=linha_conteudo).font = Font(size=10)
            linha_atual += 1
        
        linha_atual += 2  # Espaço entre seções
    
    # Ajustar largura da coluna
    ws3.column_dimensions['A'].width = 100
    
    # Salvar arquivo
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    nome_arquivo = f"FIDC_Documentacao_Abas_2_3_{timestamp}.xlsx"
    
    try:
        wb.save(nome_arquivo)
        print(f"✅ Arquivo Excel criado com sucesso!")
        print(f"📁 Nome: {nome_arquivo}")
        print(f"📋 Contém:")
        print(f"   • Aba 2: Dicionário completo das 42 colunas")
        print(f"   • Aba 3: Metodologia detalhada dos cálculos")
        print(f"🎯 Arquivo pronto para uso!")
        return nome_arquivo
    except Exception as e:
        print(f"❌ Erro ao salvar arquivo: {e}")
        return None

if __name__ == "__main__":
    criar_excel_documentacao()
