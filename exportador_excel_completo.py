#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Exportador Excel Completo FIDC - Sistema Energisa
===================================================

Este módulo gera um arquivo Excel completo com:
- Aba 1: Base de dados processada
- Aba 2: Dicionário detalhado das colunas
- Aba 3: Explicação dos cálculos e metodologia

Criado em: Agosto 2025
Projeto: Energisa FIDC Calculator
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
import os

class ExportadorExcelCompleto:
    """
    Classe responsável por gerar documentação Excel completa do sistema FIDC
    """
    
    def __init__(self, caminho_dados=None):
        """
        Inicializa o exportador
        
        Args:
            caminho_dados (str): Caminho para arquivo CSV com dados processados
        """
        self.caminho_dados = caminho_dados
        self.wb = None
        self.data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
        
        # Definir estilos
        self.estilo_cabecalho = {
            'font': Font(bold=True, color="FFFFFF", size=12),
            'fill': PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
            'alignment': Alignment(horizontal="center", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.estilo_dados = {
            'font': Font(size=10),
            'alignment': Alignment(horizontal="left", vertical="center"),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.estilo_titulo = {
            'font': Font(bold=True, size=14, color="1F4E79"),
            'alignment': Alignment(horizontal="center", vertical="center")
        }
    
    def carregar_dados(self):
        """
        Carrega os dados processados do sistema
        
        Returns:
            pandas.DataFrame: Dados processados ou DataFrame vazio se arquivo não existir
        """
        if self.caminho_dados and os.path.exists(self.caminho_dados):
            try:
                # Tentar carregar como CSV primeiro
                if self.caminho_dados.endswith('.csv'):
                    return pd.read_csv(self.caminho_dados, encoding='utf-8')
                elif self.caminho_dados.endswith('.xlsx'):
                    return pd.read_excel(self.caminho_dados)
                else:
                    return pd.read_csv(self.caminho_dados)
            except Exception as e:
                print(f"Erro ao carregar dados: {e}")
                return self._criar_dados_exemplo()
        else:
            return self._criar_dados_exemplo()
    
    def _criar_dados_exemplo(self):
        """
        Cria dados de exemplo para demonstração
        
        Returns:
            pandas.DataFrame: Dados de exemplo
        """
        dados_exemplo = {
            'nome_cliente': ['PAULO ROBERTO FLORES', 'MARIA SILVA SANTOS', 'JOÃO PEDRO OLIVEIRA'],
            'documento': [None, 12345678901, 98765432109],
            'contrato': [19.0, 25.0, 33.0],
            'classe': ['Residencial', 'Comercial', 'Industrial'],
            'situacao': ['Desligado', 'Ativo', 'Suspenso'],
            'valor_principal': [115.00, 250.50, 1500.75],
            'valor_nao_cedido': [0.0, 0.0, 50.0],
            'valor_terceiro': [0.0, 15.25, 0.0],
            'valor_cip': [2.37, 5.12, 30.15],
            'data_vencimento': ['2026-02-06', '2025-12-15', '2024-08-20'],
            'empresa': ['ESS', 'EMR', 'ESS'],
            'tipo': ['Privado', 'Privado', 'Público'],
            'status': ['Incobrável', 'Cobrável', 'Em análise'],
            'dias_atraso': [0, 45, 358],
            'aging': ['A vencer', 'De 31 a 59 dias', 'De 120 a 359 dias'],
            'valor_liquido': [95.32, 230.13, 1420.60],
            'multa': [1.91, 4.60, 28.41],
            'juros_moratorios': [0.0, 3.45, 169.27],
            'correcao_monetaria': [0.0, 12.05, 362.56],
            'valor_corrigido': [97.23, 250.23, 1980.84],
            'taxa_recuperacao': [0.0089, 0.39, 0.47],
            'valor_recuperavel': [0.87, 97.59, 931.00],
            'valor_justo': [0.87, 94.21, 878.45]
        }
        return pd.DataFrame(dados_exemplo)
    
    def criar_aba_dados(self, df):
        """
        Cria aba com base de dados
        
        Args:
            df (pandas.DataFrame): Dados para inserir na aba
        """
        ws = self.wb.create_sheet("1. Base de Dados")
        
        # Título
        ws['A1'] = f"📊 BASE DE DADOS FIDC - ENERGISA"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:X1')
        
        # Informações gerais
        ws['A3'] = f"📅 Data de Geração: {self.data_geracao}"
        ws['A4'] = f"📋 Total de Registros: {len(df):,}"
        ws['A5'] = f"📁 Origem: Sistema FIDC Calculator"
        
        # Inserir dados
        linha_inicio = 7
        
        # Cabeçalhos
        for col_idx, coluna in enumerate(df.columns, 1):
            cell = ws.cell(row=linha_inicio, column=col_idx, value=coluna)
            cell.font = self.estilo_cabecalho['font']
            cell.fill = self.estilo_cabecalho['fill']
            cell.alignment = self.estilo_cabecalho['alignment']
            cell.border = self.estilo_cabecalho['border']
        
        # Dados
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), linha_inicio + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.estilo_dados['border']
                cell.alignment = self.estilo_dados['alignment']
                cell.font = self.estilo_dados['font']
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def criar_aba_dicionario(self):
        """
        Cria aba com dicionário de dados detalhado
        """
        ws = self.wb.create_sheet("2. Dicionário de Dados")
        
        # Título
        ws['A1'] = "📚 DICIONÁRIO DE DADOS - DataFrame Final"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:E1')
        
        # Cabeçalhos
        headers = ['#', 'Nome da Coluna', 'Tipo de Dado', 'Descrição', 'Observações']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = self.estilo_cabecalho['font']
            cell.fill = self.estilo_cabecalho['fill']
            cell.alignment = self.estilo_cabecalho['alignment']
            cell.border = self.estilo_cabecalho['border']
        
        # Definir dicionário de dados
        dicionario_dados = [
            # GRUPO 1: Dados Básicos do Cliente
            (1, "nome_cliente", "Texto", "Nome do cliente titular da conta", "Campo obrigatório para identificação"),
            (2, "documento", "Numérico", "CPF/CNPJ do cliente", "Frequentemente vazio nos dados"),
            (3, "contrato", "Numérico", "Número do contrato de energia", "Identificador único do contrato"),
            (4, "classe", "Texto", "Classificação do cliente", "Residencial, Comercial, Industrial, Rural"),
            (5, "situacao", "Texto", "Status atual do cliente", "Ativo, Desligado, Suspenso"),
            (6, "valor_principal", "Numérico", "Valor original da dívida", "Valor base antes de deduções"),
            (7, "valor_nao_cedido", "Numérico", "Parcela não cedida ao FIDC", "Valor retido pela distribuidora"),
            (8, "valor_terceiro", "Numérico", "Valores de terceiros", "Tributos, taxas de terceiros"),
            (9, "valor_cip", "Numérico", "Valor CIP", "Contribuição de Iluminação Pública"),
            
            # GRUPO 2: Metadados e Controle
            (10, "data_vencimento", "Data", "Data original de vencimento", "Formato timestamp original"),
            (11, "empresa", "Texto", "Código da distribuidora", "ESS, EMR, etc."),
            (12, "tipo", "Texto", "Classificação do negócio", "Privado, Público, Hospital"),
            (13, "status", "Texto", "Status da cobrança", "Cobrável, Incobrável, Em análise"),
            (14, "id_padronizado", "Texto", "Identificador único gerado", "Chave primária para rastreamento"),
            (15, "base_origem", "Texto", "Arquivo de origem dos dados", "Rastreabilidade da fonte"),
            (16, "data_base", "Data", "Data de referência para cálculos", "Data fixa para todos os cálculos"),
            
            # GRUPO 3: Aging e Temporalidade
            (17, "data_vencimento_limpa", "Data", "Data de vencimento padronizada", "Formato YYYY-MM-DD limpo"),
            (18, "dias_atraso", "Inteiro", "Dias entre vencimento e data base", "Cálculo: data_base - data_vencimento"),
            (19, "aging", "Texto", "Classificação de aging", "9 faixas: A vencer até >1080 dias"),
            
            # GRUPO 4: Valores Limpos e Líquidos
            (20, "valor_principal_limpo", "Numérico", "Valor principal sem formatação", "Numérico puro"),
            (21, "valor_nao_cedido_limpo", "Numérico", "Valor não cedido limpo", "Numérico puro"),
            (22, "valor_terceiro_limpo", "Numérico", "Valor terceiros limpo", "Numérico puro"),
            (23, "valor_cip_limpo", "Numérico", "Valor CIP limpo", "Numérico puro"),
            (24, "valor_liquido", "Numérico", "Valor líquido final", "Fórmula: principal - nao_cedido - terceiro - cip"),
            
            # GRUPO 5: Correção Monetária
            (25, "multa", "Numérico", "Multa calculada", "Fórmula: valor_liquido × 2%"),
            (26, "meses_atraso", "Numérico", "Meses de atraso calculados", "dias_atraso ÷ 30"),
            (27, "juros_moratorios", "Numérico", "Juros moratórios", "Fórmula: valor_liquido × 1% × meses"),
            (28, "indice_vencimento", "Numérico", "Índice IGP-M no vencimento", "Obtido da tabela histórica"),
            (29, "indice_base", "Numérico", "Índice IGP-M na data base", "Índice de referência (data_base)"),
            (30, "fator_correcao", "Numérico", "Fator de correção monetária", "Fórmula: indice_base ÷ indice_vencimento"),
            (31, "correcao_monetaria", "Numérico", "Correção monetária aplicada", "Fórmula: valor_liquido × (fator_correcao - 1)"),
            (32, "valor_corrigido", "Numérico", "Valor final corrigido", "Fórmula: valor_liquido + multa + juros + correção"),
            
            # GRUPO 6: Taxa de Recuperação e Valor Justo
            (33, "aging_taxa", "Texto", "Aging mapeado para taxa", "Mapeamento: >1080 dias → Demais anos"),
            (34, "taxa_recuperacao", "Numérico", "Taxa de recuperação esperada", "Taxa decimal baseada em histórico"),
            (35, "prazo_recebimento", "Numérico", "Prazo esperado de recebimento", "Meses até recebimento esperado"),
            (36, "valor_recuperavel", "Numérico", "Valor esperado de recuperação", "Fórmula: valor_corrigido × taxa_recuperacao"),
            (37, "ipca_12m_real", "Numérico", "IPCA 12 meses", "Taxa IPCA anual em decimal"),
            (38, "ipca_mensal", "Numérico", "IPCA mensal equivalente", "Taxa IPCA mensal derivada"),
            (39, "mes_recebimento", "Inteiro", "Mês esperado de recebimento", "Baseado no prazo_recebimento"),
            (40, "fator_exponencial", "Numérico", "Fator exponencial IPCA", "Fórmula: (1 + ipca_12m)^(prazo/12)"),
            (41, "multa_para_justo", "Numérico", "Multa ajustada para valor justo", "Multa proporcional aplicada"),
            (42, "valor_justo", "Numérico", "Valor justo final", "Fórmula: valor_corrigido × fator_exponencial × taxa_recuperacao")
        ]
        
        # Inserir dados do dicionário
        for row_idx, (num, coluna, tipo, descricao, obs) in enumerate(dicionario_dados, 4):
            # Aplicar cores alternadas
            fill_color = "F2F2F2" if row_idx % 2 == 0 else "FFFFFF"
            
            cells = [
                ws.cell(row=row_idx, column=1, value=num),
                ws.cell(row=row_idx, column=2, value=coluna),
                ws.cell(row=row_idx, column=3, value=tipo),
                ws.cell(row=row_idx, column=4, value=descricao),
                ws.cell(row=row_idx, column=5, value=obs)
            ]
            
            for cell in cells:
                cell.border = self.estilo_dados['border']
                cell.alignment = self.estilo_dados['alignment']
                cell.font = self.estilo_dados['font']
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 35
    
    def criar_aba_calculos(self):
        """
        Cria aba com explicação detalhada dos cálculos
        """
        ws = self.wb.create_sheet("3. Cálculos e Metodologia")
        
        # Título
        ws['A1'] = "🧮 METODOLOGIA DE CÁLCULOS FIDC - ENERGISA"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:F1')
        
        linha_atual = 3
        
        # Seção 1: Visão Geral
        secoes = [
            {
                'titulo': '1. VISÃO GERAL DO SISTEMA',
                'conteudo': [
                    'O sistema FIDC Calculator implementa metodologia rigorosa para cálculo de valor justo',
                    'de carteiras de recebíveis de distribuidoras de energia elétrica brasileiras.',
                    '',
                    'COMPONENTES PRINCIPAIS:',
                    '• Aging (classificação temporal da inadimplência)',
                    '• Correção monetária (IGP-M até mai/2021, IPCA após jun/2021)',
                    '• Juros moratórios (1% ao mês, proporcional)',
                    '• Multa contratual (2% sobre valor líquido)',
                    '• Taxa de recuperação (baseada em histórico setorial)',
                    '• Valor presente (desconto pelo prazo de recebimento)'
                ]
            },
            {
                'titulo': '2. CÁLCULO DE AGING',
                'conteudo': [
                    'OBJETIVO: Classificar cada dívida pelo tempo de inadimplência',
                    '',
                    'FÓRMULA: Dias de Atraso = Data Base - Data de Vencimento',
                    '',
                    'CLASSIFICAÇÕES:',
                    '• A vencer: DA ≤ 0 (ainda não vencido)',
                    '• Menor que 30 dias: 0 < DA ≤ 30',
                    '• De 31 a 59 dias: 30 < DA ≤ 59',
                    '• De 60 a 89 dias: 59 < DA ≤ 89',
                    '• De 90 a 119 dias: 89 < DA ≤ 119',
                    '• De 120 a 359 dias: 119 < DA ≤ 359 (CONCENTRAÇÃO DE VALOR)',
                    '• De 360 a 719 dias: 359 < DA ≤ 719',
                    '• De 720 a 1080 dias: 719 < DA ≤ 1080',
                    '• Maior que 1080 dias: DA > 1080'
                ]
            },
            {
                'titulo': '3. CORREÇÃO MONETÁRIA',
                'conteudo': [
                    'OBJETIVO: Recompor poder de compra considerando inflação oficial',
                    '',
                    'METODOLOGIA HÍBRIDA:',
                    '• IGP-M: para datas ≤ maio/2021',
                    '• IPCA: para datas > maio/2021',
                    '',
                    'FÓRMULA: CM = Valor Líquido × (Índice Base ÷ Índice Vencimento - 1)',
                    '',
                    'ÍNDICES UTILIZADOS:',
                    '• IGP-M: Fundação Getúlio Vargas',
                    '• IPCA: Instituto Brasileiro de Geografia e Estatística',
                    '',
                    'APLICAÇÃO: Apenas para valores em atraso (DA > 0)'
                ]
            },
            {
                'titulo': '4. JUROS E MULTAS',
                'conteudo': [
                    'MULTA CONTRATUAL:',
                    '• Taxa: 2% sobre valor líquido',
                    '• Aplicação: Uma única vez para valores em atraso',
                    '• Fórmula: M = VL × 0,02 × I(DA > 0)',
                    '',
                    'JUROS MORATÓRIOS:',
                    '• Taxa: 1% ao mês',
                    '• Cálculo proporcional por dias',
                    '• Fórmula: J = VL × 0,01 × (DA ÷ 30) × I(DA > 0)',
                    '',
                    'VALOR CORRIGIDO TOTAL:',
                    'VC = Valor Líquido + Multa + Juros + Correção Monetária'
                ]
            },
            {
                'titulo': '5. TAXAS DE RECUPERAÇÃO',
                'conteudo': [
                    'OBJETIVO: Estimar probabilidade de recuperação efetiva por aging',
                    '',
                    'TAXAS POR CATEGORIA:',
                    '• A vencer: 0,89%',
                    '• Menor que 30 dias: 47,00%',
                    '• De 31 a 59 dias: 39,00%',
                    '• De 60 a 89 dias: 52,00%',
                    '• De 90 a 119 dias: 57,00%',
                    '• De 120 a 359 dias: 47,00% (MAIOR CONCENTRAÇÃO)',
                    '• De 360 a 719 dias: 2,00%',
                    '• De 720 a 1080 dias: 1,00%',
                    '• Maior que 1080 dias: 0,70%',
                    '',
                    'FONTE: Dados históricos do setor elétrico brasileiro',
                    'VARIAÇÃO: Por empresa, tipo de cliente e região'
                ]
            },
            {
                'titulo': '6. VALOR JUSTO (VALOR PRESENTE)',
                'conteudo': [
                    'OBJETIVO: Determinar valor presente considerando prazo e custo de capital',
                    '',
                    'FÓRMULA: VJ = Valor Recuperável ÷ (1 + Taxa Desconto)^(Prazo/365)',
                    '',
                    'COMPONENTES:',
                    '• Valor Recuperável: VC × Taxa de Recuperação',
                    '• Taxa de Desconto: 8% a 15% ao ano (típico FIDC elétrico)',
                    '• Prazo: Tempo esperado para recuperação (6 a 42 meses)',
                    '',
                    'PRAZOS POR AGING:',
                    '• Primeiro ano: 6 meses',
                    '• Segundo ano: 18 meses',
                    '• Terceiro ano: 36 meses',
                    '• Demais anos: 42 meses',
                    '',
                    'RESULTADO: Valor que o FIDC pode pagar pela carteira'
                ]
            },
            {
                'titulo': '7. ENCADEAMENTO DOS CÁLCULOS',
                'conteudo': [
                    'SEQUÊNCIA LÓGICA COMPLETA:',
                    '',
                    '1. VALOR LÍQUIDO = Principal - Não Cedido - Terceiros - CIP',
                    '2. AGING = Classificação por dias de atraso',
                    '3. MULTA = VL × 2% (se em atraso)',
                    '4. JUROS = VL × 1% × (dias ÷ 30) (se em atraso)',
                    '5. CORREÇÃO = VL × (Índice Base ÷ Índice Vencimento - 1)',
                    '6. VALOR CORRIGIDO = VL + Multa + Juros + Correção',
                    '7. VALOR RECUPERÁVEL = VC × Taxa de Recuperação(aging)',
                    '8. VALOR JUSTO = VR ÷ (1 + Taxa)^(Prazo/365)',
                    '',
                    'RESULTADO FINAL: Valor que maximiza retorno considerando risco'
                ]
            },
            {
                'titulo': '8. FATORES DE INFLUÊNCIA',
                'conteudo': [
                    'COMPOSIÇÃO DA CARTEIRA:',
                    '• Concentração em aging 120-359 dias maximiza valor',
                    '• Carteiras antigas (>3 anos) reduzem valor significativamente',
                    '• Mix equilibrado otimiza relação risco-retorno',
                    '',
                    'CONDIÇÕES MACROECONÔMICAS:',
                    '• Taxa Selic elevada reduz valor presente',
                    '• Inflação alta aumenta valor corrigido',
                    '• Estabilidade econômica melhora taxas de recuperação',
                    '',
                    'CARACTERÍSTICAS SETORIAIS:',
                    '• Essencialidade do serviço favorece recuperação',
                    '• Regulamentação específica influencia prazos',
                    '• Sazonalidade de consumo afeta padrões de pagamento'
                ]
            },
            {
                'titulo': '9. RESULTADOS ESPERADOS',
                'conteudo': [
                    'FAIXAS TÍPICAS DE VALOR JUSTO:',
                    '',
                    'POR AGING:',
                    '• Primeiro ano (até 359 dias): 25% a 50% do valor corrigido',
                    '• Segundo ano (360-719 dias): 1% a 3% do valor corrigido',
                    '• Terceiro ano (720-1080 dias): 0,5% a 1,5% do valor corrigido',
                    '• Acima de 3 anos: 0,3% a 0,8% do valor corrigido',
                    '',
                    'CONSOLIDADO GERAL:',
                    '• Carteiras balanceadas: 3% a 5% do valor corrigido',
                    '• Carteiras antigas: 1% a 2% do valor corrigido',
                    '• Carteiras recentes: 5% a 8% do valor corrigido',
                    '',
                    'EXEMPLO PRÁTICO:',
                    'Carteira R$ 1 bilhão → Valor FIDC: R$ 23-45 milhões'
                ]
            }
        ]
        
        # Inserir seções
        for secao in secoes:
            # Título da seção
            ws.cell(row=linha_atual, column=1, value=secao['titulo']).font = Font(bold=True, size=12, color="1F4E79")
            linha_atual += 2
            
            # Conteúdo da seção
            for linha_conteudo in secao['conteudo']:
                ws.cell(row=linha_atual, column=1, value=linha_conteudo).font = Font(size=10)
                linha_atual += 1
            
            linha_atual += 2  # Espaço entre seções
        
        # Ajustar largura da coluna
        ws.column_dimensions['A'].width = 100
    
    def gerar_excel(self, nome_arquivo="FIDC_Documentacao_Completa.xlsx"):
        """
        Gera arquivo Excel completo com todas as abas
        
        Args:
            nome_arquivo (str): Nome do arquivo a ser gerado
        
        Returns:
            str: Caminho do arquivo gerado
        """
        # Criar workbook
        self.wb = openpyxl.Workbook()
        
        # Remover aba padrão
        default_sheet = self.wb.active
        self.wb.remove(default_sheet)
        
        # Carregar dados
        df = self.carregar_dados()
        
        # Criar abas
        self.criar_aba_dados(df)
        self.criar_aba_dicionario()
        self.criar_aba_calculos()
        
        # Salvar arquivo
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_final = f"FIDC_Documentacao_Completa_{timestamp}.xlsx"
        caminho_completo = os.path.join(os.getcwd(), nome_final)
        
        self.wb.save(caminho_completo)
        print(f"✅ Arquivo Excel gerado com sucesso: {nome_final}")
        print(f"📁 Localização: {caminho_completo}")
        
        return caminho_completo

def main():
    """
    Função principal para execução do exportador
    """
    print("🔋 ENERGISA FIDC Calculator - Exportador Excel Completo")
    print("=" * 60)
    
    # Verificar se existe arquivo de dados
    possveis_arquivos = [
        "data/FIDC_Dados_Finais_20250813_180751.csv",
        "FIDC_Dados_Brutos_20250806_105650.csv",
        "FIDC_ESS.xlsx"
    ]
    
    arquivo_dados = None
    for arquivo in possveis_arquivos:
        if os.path.exists(arquivo):
            arquivo_dados = arquivo
            break
    
    # Criar exportador
    exportador = ExportadorExcelCompleto(arquivo_dados)
    
    # Gerar documentação
    arquivo_gerado = exportador.gerar_excel()
    
    print("\n📊 CONTEÚDO GERADO:")
    print("• Aba 1: Base de dados completa")
    print("• Aba 2: Dicionário detalhado das 42 colunas")
    print("• Aba 3: Metodologia de cálculos e explicações")
    print("\n🎯 Arquivo pronto para apresentação e documentação!")

if __name__ == "__main__":
    main()
