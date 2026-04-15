"""
FIDC Calculator v2 — Engine Vetorizada
======================================
Motor de cálculo puro (sem Streamlit, sem IO).
100% vetorizado: pd.cut / merge_asof / numpy broadcasting.

Fórmulas idênticas à vw_fidc_results (Supabase) e ao calculator_vectorized.py
do worker Railway — fonte da verdade confirmada.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════════════════
# TABELAS DE REFERÊNCIA
# ══════════════════════════════════════════════════════════════════════

AGING_BINS: list[float] = [-np.inf, 0, 30, 59, 89, 119, 359, 719, 1080, np.inf]
AGING_LABELS: list[str] = [
    "A vencer", "Menor que 30 dias", "De 31 a 59 dias",
    "De 60 a 89 dias", "De 90 a 119 dias", "De 120 a 359 dias",
    "De 360 a 719 dias", "De 720 a 1080 dias", "Maior que 1080 dias",
]

AGING_TAXA_MAP: dict[str, str] = {
    "A vencer": "A vencer",
    "Menor que 30 dias": "Primeiro ano",
    "De 31 a 59 dias": "Primeiro ano",
    "De 60 a 89 dias": "Primeiro ano",
    "De 90 a 119 dias": "Primeiro ano",
    "De 120 a 359 dias": "Primeiro ano",
    "De 360 a 719 dias": "Segundo ano",
    "De 720 a 1080 dias": "Terceiro ano",
    "Maior que 1080 dias": "Demais anos",
}

REMUNERATION_DISCOUNTS: dict[str, float] = {
    "A vencer": 0.065,
    "Menor que 30 dias": 0.065,
    "De 31 a 59 dias": 0.065,
    "De 60 a 89 dias": 0.065,
    "De 90 a 119 dias": 0.08,
    "De 120 a 359 dias": 0.15,
    "De 360 a 719 dias": 0.22,
    "De 720 a 1080 dias": 0.36,
    "Maior que 1080 dias": 0.50,
}

VOLTZ_DEFAULT_RATES: dict[str, dict] = {
    "A vencer":     {"taxa": 0.98, "prazo": 6},
    "Primeiro ano": {"taxa": 0.90, "prazo": 6},
    "Segundo ano":  {"taxa": 0.75, "prazo": 12},
    "Terceiro ano": {"taxa": 0.60, "prazo": 18},
    "Quarto ano":   {"taxa": 0.45, "prazo": 24},
    "Quinto ano":   {"taxa": 0.30, "prazo": 30},
    "Demais anos":  {"taxa": 0.15, "prazo": 36},
}

# Mapeamento de sinônimos de colunas → campo interno
COL_SYNONYMS: dict[str, list[str]] = {
    "valor_principal": [
        "valor_principal", "valor principal", "vp", "vlr principal",
        "saldo principal", "principal", "saldo_devedor", "valor devedor",
        "vlr_principal", "saldo_principal",
    ],
    "valor_nao_cedido": [
        "valor_nao_cedido", "valor não cedido", "valor nao cedido",
        "vlr não cedido", "vlr nao cedido", "nao cedido", "não cedido", "vnc",
    ],
    "valor_terceiro": [
        "valor_terceiro", "valor terceiro", "vlr terceiro",
        "valor_terceiros", "valor terceiros", "vt",
    ],
    "valor_cip": [
        "valor_cip", "valor cip", "vlr cip", "cip",
    ],
    "data_vencimento": [
        "data_vencimento", "data vencimento", "dt vencimento",
        "vencimento", "data de vencimento", "data_vcto", "vcto",
    ],
    "data_base": [
        "data_base", "data base", "dt base", "data de referência",
        "data referencia", "data ref", "referencia",
    ],
    "empresa": [
        "empresa", "distribuidora", "concessionaria",
        "cedente", "empresa_cedente",
    ],
    "tipo": [
        "tipo", "tipo_titulo", "tipo titulo", "categoria", "produto",
    ],
    "nome_cliente": [
        "nome_cliente", "nome cliente", "nome", "cliente", "devedor",
    ],
    "documento": [
        "documento", "cpf", "cnpj", "cpf_cnpj", "cpf/cnpj", "doc",
    ],
    "contrato": [
        "contrato", "nr_contrato", "numero contrato", "num_contrato",
    ],
    "classe": ["classe", "classe_titulo"],
    "situacao": ["situacao", "situação", "status_titulo"],
    "status_conta": ["status_conta", "status conta", "status"],
}

REQUIRED_COLS  = ["valor_principal", "data_vencimento"]
IMPORTANT_COLS = ["empresa", "data_base"]

# ══════════════════════════════════════════════════════════════════════
# ÍNDICE IGP-M HISTÓRICO (hardcoded, base ago/1994=100)
# ══════════════════════════════════════════════════════════════════════
# Fonte: FGV — mesmo dicionário de parametros_correcao.py original

IGPM_DICT: dict[str, float] = {
    "1994.08": 100.00000, "1994.09": 101.75100, "1994.10": 103.60200,
    "1994.11": 106.55300, "1994.12": 107.45000, "1995.01": 108.44200,
    "1995.02": 109.94500, "1995.03": 111.17800, "1995.04": 113.51800,
    "1995.05": 114.17100, "1995.06": 116.98400, "1995.07": 119.11400,
    "1995.08": 121.72900, "1995.09": 120.86900, "1995.10": 121.50300,
    "1995.11": 122.95500, "1995.12": 123.83300, "1996.01": 125.97700,
    "1996.02": 127.20200, "1996.03": 127.71500, "1996.04": 128.13000,
    "1996.05": 130.12100, "1996.06": 131.44500, "1996.07": 133.21300,
    "1996.08": 133.58700, "1996.09": 133.72200, "1996.10": 133.97800,
    "1996.11": 134.24200, "1996.12": 135.22500, "1997.01": 137.61300,
    "1997.02": 138.20400, "1997.03": 139.79500, "1997.04": 140.74200,
    "1997.05": 141.04000, "1997.06": 142.09000, "1997.07": 142.22100,
    "1997.08": 142.35300, "1997.09": 143.04200, "1997.10": 143.56700,
    "1997.11": 144.48100, "1997.12": 145.69500, "1998.01": 147.09100,
    "1998.02": 147.35600, "1998.03": 147.63500, "1998.04": 147.82100,
    "1998.05": 148.02100, "1998.06": 148.58800, "1998.07": 148.33900,
    "1998.08": 148.10900, "1998.09": 147.98400, "1998.10": 148.10000,
    "1998.11": 147.62800, "1998.12": 148.29100, "1999.01": 149.53300,
    "1999.02": 154.93300, "1999.03": 159.32500, "1999.04": 160.45900,
    "1999.05": 159.99600, "1999.06": 160.57300, "1999.07": 163.06000,
    "1999.08": 165.60300, "1999.09": 167.99700, "1999.10": 170.86100,
    "1999.11": 174.93900, "1999.12": 178.09900, "2000.01": 180.30100,
    "2000.02": 180.93500, "2000.03": 181.21400, "2000.04": 181.63500,
    "2000.05": 182.18900, "2000.06": 183.74500, "2000.07": 186.63400,
    "2000.08": 191.08700, "2000.09": 193.29700, "2000.10": 194.04000,
    "2000.11": 194.59900, "2000.12": 195.82700, "2001.01": 197.04500,
    "2001.02": 197.49100, "2001.03": 198.60600, "2001.04": 200.59100,
    "2001.05": 202.32400, "2001.06": 204.31000, "2001.07": 207.34100,
    "2001.08": 210.21100, "2001.09": 210.85300, "2001.10": 213.33900,
    "2001.11": 215.68500, "2001.12": 216.16300, "2002.01": 216.94400,
    "2002.02": 217.07400, "2002.03": 217.27600, "2002.04": 218.48600,
    "2002.05": 220.29200, "2002.06": 223.68800, "2002.07": 228.05700,
    "2002.08": 233.34800, "2002.09": 238.94300, "2002.10": 248.19900,
    "2002.11": 261.08000, "2002.12": 270.86700, "2003.01": 277.17300,
    "2003.02": 283.50600, "2003.03": 287.85500, "2003.04": 290.51200,
    "2003.05": 289.74700, "2003.06": 286.84300, "2003.07": 285.64900,
    "2003.08": 286.73500, "2003.09": 290.12700, "2003.10": 291.22900,
    "2003.11": 292.65700, "2003.12": 294.45500, "2004.01": 297.03900,
    "2004.02": 299.09700, "2004.03": 302.48400, "2004.04": 306.15100,
    "2004.05": 310.15200, "2004.06": 314.41900, "2004.07": 318.53200,
    "2004.08": 322.41200, "2004.09": 324.65100, "2004.10": 325.92500,
    "2004.11": 328.58800, "2004.12": 331.00500, "2005.01": 332.29800,
    "2005.02": 333.28800, "2005.03": 336.12300, "2005.04": 339.03000,
    "2005.05": 338.29900, "2005.06": 336.80100, "2005.07": 335.66300,
    "2005.08": 333.47400, "2005.09": 331.69000, "2005.10": 333.69400,
    "2005.11": 335.03300, "2005.12": 335.00600, "2006.01": 338.08300,
    "2006.02": 338.12800, "2006.03": 337.33900, "2006.04": 335.92100,
    "2006.05": 337.18500, "2006.06": 339.71200, "2006.07": 340.31200,
    "2006.08": 341.57400, "2006.09": 342.56100, "2006.10": 344.15500,
    "2006.11": 346.74600, "2006.12": 347.84200, "2007.01": 349.59300,
    "2007.02": 350.52400, "2007.03": 351.71700, "2007.04": 351.86900,
    "2007.05": 352.02000, "2007.06": 352.93600, "2007.07": 353.92000,
    "2007.08": 357.40400, "2007.09": 361.99700, "2007.10": 365.79400,
    "2007.11": 368.33400, "2007.12": 374.81500, "2008.01": 378.90000,
    "2008.02": 380.90600, "2008.03": 383.73100, "2008.04": 386.38000,
    "2008.05": 392.59200, "2008.06": 400.38200, "2008.07": 407.44600,
    "2008.08": 406.12700, "2008.09": 406.55700, "2008.10": 410.52400,
    "2008.11": 412.10400, "2008.12": 411.57500, "2009.01": 409.78200,
    "2009.02": 410.84900, "2009.03": 407.80800, "2009.04": 407.18100,
    "2009.05": 406.88500, "2009.06": 406.48600, "2009.07": 404.71800,
    "2009.08": 403.25300, "2009.09": 404.94500, "2009.10": 405.12900,
    "2009.11": 405.54800, "2009.12": 404.49900, "2010.01": 407.04900,
    "2010.02": 411.84300, "2010.03": 415.73400, "2010.04": 418.91700,
    "2010.05": 423.88500, "2010.06": 427.48900, "2010.07": 428.15000,
    "2010.08": 431.44500, "2010.09": 436.42300, "2010.10": 440.82900,
    "2010.11": 447.20600, "2010.12": 450.30100, "2011.01": 453.87500,
    "2011.02": 458.39700, "2011.03": 461.24900, "2011.04": 463.31100,
    "2011.05": 465.31100, "2011.06": 464.46300, "2011.07": 463.92700,
    "2011.08": 465.96800, "2011.09": 468.97500, "2011.10": 471.46600,
    "2011.11": 473.80800, "2011.12": 473.25200, "2012.01": 474.42900,
    "2012.02": 474.13800, "2012.03": 476.16600, "2012.04": 480.22900,
    "2012.05": 485.14000, "2012.06": 488.34200, "2012.07": 494.89100,
    "2012.08": 501.95700, "2012.09": 506.80400, "2012.10": 506.92600,
    "2012.11": 506.79500, "2012.12": 510.25200, "2013.01": 511.97700,
    "2013.02": 513.46700, "2013.03": 514.52600, "2013.04": 515.27600,
    "2013.05": 515.29900, "2013.06": 519.15300, "2013.07": 520.50800,
    "2013.08": 521.27000, "2013.09": 529.08500, "2013.10": 533.62100,
    "2013.11": 535.16800, "2013.12": 538.37000, "2014.01": 540.95900,
    "2014.02": 543.03800, "2014.03": 552.08700, "2014.04": 556.42000,
    "2014.05": 555.67900, "2014.06": 551.55400, "2014.07": 548.20200,
    "2014.08": 546.74500, "2014.09": 547.83900, "2014.10": 549.39600,
    "2014.11": 554.76900, "2014.12": 558.21300, "2015.01": 562.48200,
    "2015.02": 564.00400, "2015.03": 569.53600, "2015.04": 576.17500,
    "2015.05": 578.51600, "2015.06": 582.40100, "2015.07": 586.42600,
    "2015.08": 588.04200, "2015.09": 593.60600, "2015.10": 604.83200,
    "2015.11": 614.05100, "2015.12": 617.04400, "2016.01": 624.06000,
    "2016.02": 632.11400, "2016.03": 635.34900, "2016.04": 637.43400,
    "2016.05": 642.65100, "2016.06": 653.49600, "2016.07": 654.64100,
    "2016.08": 655.60200, "2016.09": 656.89400, "2016.10": 657.92700,
    "2016.11": 657.75200, "2016.12": 661.30400, "2017.01": 665.54200,
    "2017.02": 666.09900, "2017.03": 666.19700, "2017.04": 658.89800,
    "2017.05": 652.75800, "2017.06": 648.40900, "2017.07": 643.76600,
    "2017.08": 644.38300, "2017.09": 647.40000, "2017.10": 648.67200,
    "2017.11": 652.07300, "2017.12": 657.85900, "2018.01": 662.82600,
    "2018.02": 663.31100, "2018.03": 667.52400, "2018.04": 671.32700,
    "2018.05": 680.57900, "2018.06": 693.28700, "2018.07": 696.80000,
    "2018.08": 701.67700, "2018.09": 712.37300, "2018.10": 718.68400,
    "2018.11": 715.16600, "2018.12": 707.44100, "2019.01": 707.48800,
    "2019.02": 713.74700, "2019.03": 722.70700, "2019.04": 729.34600,
    "2019.05": 732.59500, "2019.06": 738.42100, "2019.07": 741.34600,
    "2019.08": 736.40200, "2019.09": 736.36200, "2019.10": 741.33300,
    "2019.11": 743.55800, "2019.12": 759.11200, "2020.01": 762.73300,
    "2020.02": 762.42300, "2020.03": 771.90800, "2020.04": 778.10100,
    "2020.05": 780.28000, "2020.06": 792.42900, "2020.07": 810.08300,
    "2020.08": 832.31300, "2020.09": 868.44200, "2020.10": 896.50500,
    "2020.11": 925.88700, "2020.12": 934.75800, "2021.01": 958.84400,
    "2021.02": 983.06300, "2021.03": 1011.94800, "2021.04": 1027.21100,
    "2021.05": 1069.28900,
}

IPCA_FALLBACK: dict[str, float] = {
    "2021.06": 1069.29, "2021.07": 1078.02, "2021.08": 1086.41,
    "2021.09": 1096.87, "2021.10": 1107.98, "2021.11": 1119.53,
    "2021.12": 1130.95, "2022.01": 1142.76, "2022.02": 1154.36,
    "2022.03": 1166.03, "2022.04": 1177.75, "2022.05": 1190.70,
    "2022.06": 1202.69, "2022.07": 1210.44, "2022.08": 1209.64,
    "2022.09": 1215.85, "2022.10": 1222.74, "2022.11": 1228.50,
    "2022.12": 1234.63, "2023.01": 1241.52, "2023.02": 1248.99,
    "2023.03": 1257.24, "2023.04": 1263.73, "2023.05": 1267.85,
    "2023.06": 1268.06, "2023.07": 1268.91, "2023.08": 1270.23,
    "2023.09": 1271.96, "2023.10": 1273.51, "2023.11": 1274.15,
    "2023.12": 1276.05, "2024.01": 1279.38, "2024.02": 1283.86,
    "2024.03": 1289.02, "2024.04": 1293.40, "2024.05": 1296.85,
    "2024.06": 1300.47, "2024.07": 1304.94, "2024.08": 1308.67,
    "2024.09": 1312.83, "2024.10": 1316.78, "2024.11": 1320.84,
    "2024.12": 1325.18, "2025.01": 1329.46, "2025.02": 1334.01,
    "2025.03": 1338.60, "2025.04": 1343.22,
}

# ══════════════════════════════════════════════════════════════════════
# CARREGAMENTO DE ÍNDICES
# ══════════════════════════════════════════════════════════════════════

def load_ipca_from_sidra(valor_base: float = 1069.29) -> dict[str, float]:
    """Baixa IPCA mensal do IBGE SIDRA e retorna índice acumulado desde 2021.06."""
    try:
        import sidrapy
        df = sidrapy.get_table(
            table_code="1737",
            territorial_level="1",
            ibge_territorial_code="all",
            variable="63",
            period=f"202106-{datetime.today().strftime('%Y%m')}",
            format="pandas",
        )[["D2C", "V"]][1:]
        df["V"] = df["V"].astype(float)
        acumulado = [valor_base]
        for i in range(1, len(df)):
            acumulado.append(round(acumulado[-1] * (1 + df.loc[i, "V"] / 100), 4))
        df["acumulado"] = acumulado
        return {
            f"{str(row['D2C'])[:4]}.{str(row['D2C'])[4:]}": row["acumulado"]
            for _, row in df.iterrows()
        }
    except Exception:
        return IPCA_FALLBACK.copy()


def build_index_series(
    igpm_dict: Optional[dict[str, float]] = None,
    ipca_dict: Optional[dict[str, float]] = None,
    df_excel: Optional[pd.DataFrame] = None,
) -> pd.DataFrame:
    """
    Constrói DataFrame de índices pronto para merge_asof.

    Prioridade:
      1. df_excel (se fornecido) — índices carregados do Excel do usuário
      2. igpm_dict + ipca_dict — séries hardcoded/SIDRA
    """
    if df_excel is not None and not df_excel.empty:
        df = df_excel[["data", "indice"]].rename(columns={"indice": "value"}).copy()
        df["date"] = pd.to_datetime(df["data"], errors="coerce")
        df = df.dropna(subset=["date"]).sort_values("date").reset_index(drop=True)
        df["value"] = pd.to_numeric(df["value"], errors="coerce").fillna(0)
        return df[["date", "value"]]

    combined = {**(igpm_dict or IGPM_DICT), **(ipca_dict or IPCA_FALLBACK)}
    records = []
    for key, val in combined.items():
        try:
            y, m = key.split(".")
            records.append({"date": pd.Timestamp(f"{y}-{m}-01"), "value": float(val)})
        except Exception:
            pass
    df = pd.DataFrame(records).sort_values("date").reset_index(drop=True)
    return df


def load_indices_from_excel(
    file_obj, sheet_name: Optional[str] = None
) -> Optional[pd.DataFrame]:
    """
    Carrega série de índices de arquivo Excel.
    Suporta duas estruturas:
      - Aba IGPM_IPCA: colunas C (ano), D (mês), F (índice)
      - Aba IGPM / genérica: colunas A (data), B (índice)
    Retorna DataFrame com colunas ['data', 'indice'] ou None em erro.
    """
    try:
        from openpyxl import load_workbook

        if hasattr(file_obj, "read"):
            if hasattr(file_obj, "seek"):
                file_obj.seek(0)
            raw = io.BytesIO(file_obj.read())
        else:
            raw = file_obj

        wb = load_workbook(raw, data_only=True, read_only=True)
        sheet = sheet_name if (sheet_name and sheet_name in wb.sheetnames) else wb.sheetnames[0]
        ws = wb[sheet]

        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        df_raw = pd.DataFrame(rows).dropna(how="all").reset_index(drop=True)
        wb.close()

        # Detectar estrutura: se ≥6 colunas → IGPM_IPCA (cols C,D,F = idx 2,3,5)
        if df_raw.shape[1] >= 6:
            df = df_raw.iloc[:, [2, 3, 5]].copy()
            df.columns = ["ano", "mes", "indice"]
            df = df.dropna()
            df["ano"] = pd.to_numeric(df["ano"], errors="coerce")
            df["mes"] = pd.to_numeric(df["mes"], errors="coerce")
            df = df.dropna(subset=["ano", "mes"])
            df = df[(df["ano"] >= 1990) & (df["ano"] <= 2035)]
            df = df[(df["mes"] >= 1) & (df["mes"] <= 12)]
            df["data"] = pd.to_datetime(
                df["ano"].astype(int).astype(str) + "-"
                + df["mes"].astype(int).astype(str).str.zfill(2) + "-01"
            )
        else:
            # Estrutura simples: col A=data, col B=índice
            df = df_raw.iloc[:, [0, 1]].copy()
            df.columns = ["data_raw", "indice"]
            df = df.dropna(subset=["data_raw"])

            datas = []
            for v in df["data_raw"]:
                if isinstance(v, datetime):
                    datas.append(pd.Timestamp(v.year, v.month, 1))
                elif isinstance(v, (int, float)):
                    from datetime import timedelta
                    d = datetime(1899, 12, 30) + timedelta(days=int(v))
                    datas.append(pd.Timestamp(d.year, d.month, 1))
                else:
                    try:
                        t = pd.to_datetime(str(v), dayfirst=True, errors="coerce")
                        datas.append(t.replace(day=1) if not pd.isnull(t) else pd.NaT)
                    except Exception:
                        datas.append(pd.NaT)
            df["data"] = datas

        df["indice"] = (
            df["indice"].astype(str).str.replace(",", ".", regex=False)
        )
        df["indice"] = pd.to_numeric(df["indice"], errors="coerce")
        df = df.dropna(subset=["data", "indice"])
        df = df[["data", "indice"]].sort_values("data").reset_index(drop=True)
        return df

    except Exception as exc:
        print(f"[engine] load_indices_from_excel erro: {exc}")
        return None


def load_recovery_rates_from_excel(file_obj) -> Optional[pd.DataFrame]:
    """
    Carrega tabela de taxas de recuperação.
    Espera colunas: Empresa, Tipo, Aging, Taxa de recuperação, Prazo de recebimento
    """
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        df = pd.read_excel(file_obj, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        # Normaliza nomes de colunas
        rename: dict[str, str] = {}
        for col in df.columns:
            cl = col.lower().strip()
            if cl in ("empresa", "distribuidora", "cedente"):
                rename[col] = "Empresa"
            elif cl in ("tipo", "tipo titulo", "tipo_titulo"):
                rename[col] = "Tipo"
            elif cl in ("aging", "faixa", "faixa aging"):
                rename[col] = "Aging"
            elif "recupera" in cl and "taxa" in cl:
                rename[col] = "Taxa de recuperação"
            elif "prazo" in cl or "recebimento" in cl:
                rename[col] = "Prazo de recebimento"
        df = df.rename(columns=rename)

        required = {"Empresa", "Aging", "Taxa de recuperação", "Prazo de recebimento"}
        if not required.issubset(df.columns):
            missing = required - set(df.columns)
            raise ValueError(f"Colunas ausentes: {missing}")

        if "Tipo" not in df.columns:
            df["Tipo"] = ""

        def _parse(s):
            s = str(s).strip().replace(",", ".")
            try:
                return float(s)
            except Exception:
                return 0.0

        df["Taxa de recuperação"] = df["Taxa de recuperação"].apply(_parse)
        df["Prazo de recebimento"] = df["Prazo de recebimento"].apply(_parse).astype(int)
        return df[["Empresa", "Tipo", "Aging", "Taxa de recuperação", "Prazo de recebimento"]]

    except Exception as exc:
        print(f"[engine] load_recovery_rates_from_excel erro: {exc}")
        return None


def load_di_pre_from_excel(file_obj) -> Optional[pd.DataFrame]:
    """
    Carrega curva DI-PRE do Excel.
    Espera colunas: meses_futuros (ou Prazo), 252 (taxa 252 d.u.)
    """
    try:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        df = pd.read_excel(file_obj, dtype=str)
        df.columns = [str(c).strip() for c in df.columns]

        rename: dict[str, str] = {}
        for col in df.columns:
            cl = col.lower().strip()
            if cl in ("meses_futuros", "meses futuros", "prazo", "meses"):
                rename[col] = "meses_futuros"
            elif cl in ("252", "taxa_252", "taxa 252", "di_pre", "di pre", "taxa"):
                rename[col] = "252"
        df = df.rename(columns=rename)

        if "meses_futuros" not in df.columns or "252" not in df.columns:
            raise ValueError(f"Colunas esperadas: meses_futuros, 252. Encontradas: {list(df.columns)}")

        def _p(s):
            return float(str(s).strip().replace(",", "."))

        df["meses_futuros"] = df["meses_futuros"].apply(_p).astype(int)
        df["252"] = df["252"].apply(_p)
        return df[["meses_futuros", "252"]].dropna()

    except Exception as exc:
        print(f"[engine] load_di_pre_from_excel erro: {exc}")
        return None


# ══════════════════════════════════════════════════════════════════════
# AUTO-DETECÇÃO DE COLUNAS
# ══════════════════════════════════════════════════════════════════════

def auto_detect_columns(csv_cols: list[str]) -> dict[str, str]:
    """Detecta campos automaticamente por sinônimos. Retorna {interno: csv_col}."""
    mapping: dict[str, str] = {}
    csv_lower = {c.lower().strip(): c for c in csv_cols}
    for internal, synonyms in COL_SYNONYMS.items():
        for syn in synonyms:
            if syn in csv_lower:
                mapping[internal] = csv_lower[syn]
                break
    return mapping


# ══════════════════════════════════════════════════════════════════════
# PARSING HELPERS
# ══════════════════════════════════════════════════════════════════════

def parse_float_col(series: pd.Series) -> pd.Series:
    """Parse numérico: aceita formato pt-BR (1.234,56) e padrão."""
    s = series.astype(str).str.strip()
    result = pd.to_numeric(s, errors="coerce")
    mask = result.isna()
    if mask.any():
        br = pd.to_numeric(
            s[mask]
            .str.replace(r"\s", "", regex=True)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False),
            errors="coerce",
        )
        result = result.copy()
        result[mask] = br
    return result.fillna(0.0)


def parse_date_col(series: pd.Series) -> pd.Series:
    """Parse de data: ISO, pt-BR e genérico."""
    s = series.astype(str).str.strip()
    parsed = pd.to_datetime(s, format="%Y-%m-%d", errors="coerce")
    m = parsed.isna()
    if m.any():
        parsed = parsed.copy()
        parsed[m] = pd.to_datetime(s[m], format="%d/%m/%Y", errors="coerce")
    m2 = parsed.isna()
    if m2.any():
        parsed = parsed.copy()
        parsed[m2] = pd.to_datetime(s[m2], dayfirst=True, errors="coerce")
    return parsed


def read_uploaded_file(file_obj) -> pd.DataFrame:
    """Lê CSV ou Excel com detecção de encoding/separador."""
    name = getattr(file_obj, "name", "")
    if name.lower().endswith((".xlsx", ".xls")):
        xl = pd.ExcelFile(file_obj)
        pref = next(
            (s for s in xl.sheet_names if s.lower() in ("base", "dados", "data", "sheet1", "planilha1")),
            xl.sheet_names[0],
        )
        df = xl.parse(pref, dtype=str)
    else:
        if hasattr(file_obj, "seek"):
            file_obj.seek(0)
        raw = file_obj.read()
        for enc in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
            try:
                df = pd.read_csv(
                    io.BytesIO(raw), sep=None, engine="python",
                    encoding=enc, dtype=str, on_bad_lines="skip",
                )
                break
            except Exception:
                continue
        else:
            raise ValueError(f"Não foi possível ler: {name}")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ══════════════════════════════════════════════════════════════════════
# LOOKUP VETORIZADO DE ÍNDICES  (merge_asof — O(n log n))
# ══════════════════════════════════════════════════════════════════════

def _build_idx_df(idx_series: pd.DataFrame) -> pd.DataFrame:
    """Prepara DataFrame de índices para merge_asof."""
    df = idx_series.copy()
    if "date" not in df.columns and "data" in df.columns:
        df = df.rename(columns={"data": "date", "indice": "value"})
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["value"] = pd.to_numeric(df["value"], errors="coerce")
    return df.dropna().sort_values("date").reset_index(drop=True)


def lookup_index_values(dates: pd.Series, idx_df: pd.DataFrame) -> np.ndarray:
    """
    Para cada data em `dates`, retorna o valor mais próximo em idx_df.
    100% vetorizado via merge_asof — sem iterrows, sem apply.
    """
    dates_ts    = pd.to_datetime(dates, errors="coerce")
    nat_mask    = dates_ts.isna().values
    dates_filled = dates_ts.fillna(pd.Timestamp("2000-01-01"))

    temp = pd.DataFrame({
        "date":  dates_filled.values,
        "_pos":  np.arange(len(dates_filled), dtype=np.intp),
    }).sort_values("date")

    merged = pd.merge_asof(temp, idx_df[["date", "value"]], on="date", direction="nearest")
    result = merged.sort_values("_pos")["value"].to_numpy(dtype=float)
    result[nat_mask] = 1.0
    return result


# ══════════════════════════════════════════════════════════════════════
# TAXAS DE RECUPERAÇÃO  (merge vetorizado — 3 níveis de fallback)
# ══════════════════════════════════════════════════════════════════════

def _build_rates_df(df_taxa: pd.DataFrame) -> pd.DataFrame:
    """Padroniza colunas e tipos da tabela de taxas de recuperação."""
    if df_taxa is None or df_taxa.empty:
        return pd.DataFrame(columns=["Empresa", "Tipo", "Aging", "Taxa de recuperação", "Prazo de recebimento"])
    df = df_taxa.copy()
    df["_emp"] = df["Empresa"].astype(str).str.upper().str.strip()
    df["_tip"] = df["Tipo"].astype(str).str.lower().str.strip()
    df["Taxa de recuperação"]   = pd.to_numeric(df["Taxa de recuperação"],   errors="coerce").fillna(0.0)
    df["Prazo de recebimento"] = pd.to_numeric(df["Prazo de recebimento"], errors="coerce").fillna(6).astype(int)
    return df


def apply_recovery_rates(df: pd.DataFrame, df_taxa: Optional[pd.DataFrame]) -> pd.DataFrame:
    """
    Aplica taxas de recuperação com 3 níveis de fallback (vetorizado):
      1. empresa + tipo + aging_taxa (exato)
      2. empresa + aging_taxa (qualquer tipo)
      3. VOLTZ defaults (quando is_voltz=True)
    """
    df = df.copy()
    df["_emp"] = df["empresa"].astype(str).str.upper().str.strip()
    df["_tip"] = df["tipo"].astype(str).str.lower().str.strip()

    rates_df = _build_rates_df(df_taxa)

    if not rates_df.empty:
        # Nível 1 — match exato
        exact = rates_df[["_emp", "_tip", "Aging", "Taxa de recuperação", "Prazo de recebimento"]].rename(
            columns={"Aging": "aging_taxa", "Taxa de recuperação": "_te", "Prazo de recebimento": "_pe"}
        )
        df = df.merge(exact, on=["_emp", "_tip", "aging_taxa"], how="left")

        # Nível 2 — fallback por empresa
        fallback = (
            rates_df.sort_values(["_emp", "Aging"])
            .groupby(["_emp", "Aging"], as_index=False)
            .first()[["_emp", "Aging", "Taxa de recuperação", "Prazo de recebimento"]]
            .rename(columns={"Aging": "aging_taxa", "Taxa de recuperação": "_tf", "Prazo de recebimento": "_pf"})
        )
        df = df.merge(fallback, on=["_emp", "aging_taxa"], how="left")

        no_exact = df["_te"].isna()
        df.loc[no_exact, "_te"] = df.loc[no_exact, "_tf"]
        df.loc[no_exact, "_pe"] = df.loc[no_exact, "_pf"]
    else:
        df["_te"] = np.nan
        df["_pe"] = np.nan

    # Nível 3 — VOLTZ defaults
    voltz_df = pd.DataFrame([
        {"aging_taxa": k, "_tv": v["taxa"], "_pv": v["prazo"]}
        for k, v in VOLTZ_DEFAULT_RATES.items()
    ])
    df = df.merge(voltz_df, on="aging_taxa", how="left")

    is_voltz = df["is_voltz"].fillna(False).astype(bool)
    df["taxa_recuperacao"]  = np.where(is_voltz, df["_tv"],           df["_te"].fillna(0.0)).astype(float)
    df["prazo_recebimento"] = np.where(is_voltz, df["_pv"].fillna(6), df["_pe"].fillna(6)).astype(int)

    drop = [c for c in df.columns if c.startswith("_")]
    return df.drop(columns=drop, errors="ignore")


# ══════════════════════════════════════════════════════════════════════
# ESCALAR: IPCA MENSAL
# ══════════════════════════════════════════════════════════════════════

def get_ipca_mensal(idx_df: pd.DataFrame) -> float:
    """
    Taxa IPCA mensal real a partir dos últimos 13 pontos da série de índices.
    Fallback: 0.004 (~4.8% a.a.)
    """
    if idx_df is None or len(idx_df) < 13:
        return 0.004
    df = idx_df.sort_values("date" if "date" in idx_df.columns else "data")
    col = "value" if "value" in df.columns else "indice"
    cur  = float(df.iloc[-1][col])
    prev = float(df.iloc[-13][col])
    if prev <= 0:
        return 0.004
    return (1 + (cur / prev - 1)) ** (1 / 12) - 1


def get_di_pre_rate(df_di_pre: Optional[pd.DataFrame], meses: int) -> float:
    """Taxa DI-PRE anual (%) para o prazo em meses mais próximo. Fallback: 12%."""
    if df_di_pre is None or df_di_pre.empty:
        return 12.0
    match = df_di_pre[df_di_pre["meses_futuros"] == meses]
    if not match.empty:
        return float(match.iloc[0]["252"])
    closest = df_di_pre.iloc[(df_di_pre["meses_futuros"] - meses).abs().argsort()[:1]]
    return float(closest.iloc[0]["252"])


# ══════════════════════════════════════════════════════════════════════
# CÁLCULO PRINCIPAL VETORIZADO
# ══════════════════════════════════════════════════════════════════════

def calculate(
    df: pd.DataFrame,
    idx_df: pd.DataFrame,
    df_taxa: Optional[pd.DataFrame] = None,
    df_di_pre: Optional[pd.DataFrame] = None,
    data_base: Optional[str] = None,
    spread_percent: float = 0.025,
    prazo_horizonte: int = 6,
    is_voltz_global: bool = False,
) -> pd.DataFrame:
    """
    Pipeline de cálculo vetorizado completo.

    Fórmulas idênticas à vw_fidc_results (Supabase):
      Padrão : VL = max(VP-VNC-VT-VCIP,0)  |  JM simples  |  IGP-M/IPCA
      VOLTZ  : SDV = VP×1.0465              |  JM composto  |  IGP-M

    Retorna df com todas as colunas calculadas.
    """
    df = df.copy()

    # ── is_voltz ────────────────────────────────────────────────────────
    if is_voltz_global:
        df["is_voltz"] = True
    elif "is_voltz" not in df.columns:
        df["is_voltz"] = df.get("empresa", pd.Series(dtype=str)).str.upper().eq("VOLTZ")
    else:
        df["is_voltz"] = (
            df["is_voltz"].astype(str).str.lower().isin(["true", "1", "sim", "s"])
            | df.get("empresa", pd.Series(dtype=str)).str.upper().eq("VOLTZ")
        )
    is_voltz = df["is_voltz"].astype(bool)

    # ── data_base global ────────────────────────────────────────────────
    if "data_base" not in df.columns or df["data_base"].isna().all():
        df["data_base"] = data_base or datetime.today().strftime("%Y-%m-%d")
    else:
        df["data_base"] = parse_date_col(df["data_base"]).dt.strftime("%Y-%m-%d").where(
            parse_date_col(df["data_base"]).notna(), other=data_base or datetime.today().strftime("%Y-%m-%d")
        )

    # ── Empresa / tipo defaults ─────────────────────────────────────────
    df["empresa"] = df.get("empresa", pd.Series("DESCONHECIDA", index=df.index)).fillna("DESCONHECIDA").astype(str).str.strip()
    df["tipo"]    = df.get("tipo",    pd.Series("",             index=df.index)).fillna("").astype(str)

    # ── Escalares ───────────────────────────────────────────────────────
    idx_df_clean = _build_idx_df(idx_df)
    ipca_mensal  = get_ipca_mensal(idx_df_clean)
    di_pct       = get_di_pre_rate(df_di_pre, prazo_horizonte)
    di_anual     = di_pct / 100
    taxa_total   = (1 + di_anual) * (1 + spread_percent) - 1
    taxa_desc_mensal = (1 + taxa_total) ** (1 / 12) - 1

    # ── Datas ────────────────────────────────────────────────────────────
    dt_venc = pd.to_datetime(df["data_vencimento"], errors="coerce")
    dt_base = pd.to_datetime(df["data_base"],       errors="coerce")

    # ── Aging (pd.cut — vetorizado) ──────────────────────────────────────
    dias = (dt_base - dt_venc).dt.days.fillna(0).astype(int)
    df["dias_atraso"] = dias
    df["aging"]      = pd.cut(dias, bins=AGING_BINS, labels=AGING_LABELS, right=True).astype(str)
    df["aging_taxa"] = df["aging"].map(AGING_TAXA_MAP).fillna("Demais anos")

    is_overdue = dias > 0

    # ── Valores base ─────────────────────────────────────────────────────
    vp   = parse_float_col(df["valor_principal"])
    vnc  = parse_float_col(df.get("valor_nao_cedido", pd.Series(0, index=df.index)))
    vt   = parse_float_col(df.get("valor_terceiro",   pd.Series(0, index=df.index)))
    vcip = parse_float_col(df.get("valor_cip",        pd.Series(0, index=df.index)))

    # ── Fator de correção (merge_asof — vetorizado) ──────────────────────
    if idx_df_clean is not None and not idx_df_clean.empty and is_overdue.any():
        arr_base = lookup_index_values(dt_base, idx_df_clean)
        arr_venc = lookup_index_values(dt_venc, idx_df_clean)
        safe     = np.where(arr_venc > 0, arr_venc, 1.0)
        fator    = np.where(is_overdue, arr_base / safe, 1.0)
    else:
        fator = np.ones(len(df))
    df["fator_correcao"] = fator

    # ── Correção padrão ──────────────────────────────────────────────────
    vl_std    = np.maximum(vp - vnc - vt - vcip, 0.0)
    multa_std = np.where(is_overdue, vl_std * 0.02, 0.0)
    jm_std    = np.where(is_overdue, vl_std * 0.01 * (dias / 30.0), 0.0)
    cm_std    = np.where(is_overdue, np.maximum(vl_std * (fator - 1), 0.0), 0.0)
    vc_std    = vl_std + multa_std + jm_std + cm_std

    # ── Correção VOLTZ ────────────────────────────────────────────────────
    # SDV = VP × 1.0465  (VP + 4,65% juros remuneratórios)
    # JM  = SDV × (1,01^(dias/30,44) - 1)  [composto]
    jr_v    = vp * 0.0465
    sdv     = vp + jr_v                   # = VP × 1.0465
    cm_v    = np.where(is_overdue, np.maximum(sdv * (fator - 1), 0.0), 0.0)
    multa_v = np.where(is_overdue, sdv * 0.02, 0.0)
    jm_v    = np.where(is_overdue, sdv * ((1.01 ** (dias / 30.44)) - 1), 0.0)
    vc_v    = sdv + cm_v + multa_v + jm_v

    # ── Seleciona padrão vs VOLTZ ─────────────────────────────────────────
    df["valor_liquido"]            = np.where(is_voltz, vp,      vl_std)
    df["valor_principal_limpo"]    = vp
    df["multa"]                    = np.where(is_voltz, multa_v, multa_std)
    df["juros_moratorios"]         = np.where(is_voltz, jm_v,    jm_std)
    df["correcao_monetaria"]       = np.where(is_voltz, cm_v,    cm_std)
    df["valor_corrigido"]          = np.where(is_voltz, vc_v,    vc_std)
    df["juros_remuneratorios"]     = np.where(is_voltz, jr_v,    np.nan)
    df["saldo_devedor_vencimento"] = np.where(is_voltz, sdv,     np.nan)

    # ── Taxas de recuperação (merge vetorizado) ───────────────────────────
    df = apply_recovery_rates(df, df_taxa)

    # ── Valor Justo ────────────────────────────────────────────────────────
    # VJ = VC × TR × ((1+IPCA)^prazo + prazo×0,01) / (1+taxa_desc)^prazo × (1-desconto)
    prazo_rec = df["prazo_recebimento"].astype(float)
    taxa_rec  = df["taxa_recuperacao"].astype(float)
    vc        = df["valor_corrigido"].astype(float)

    df["valor_recuperavel"]  = vc * taxa_rec
    fc_receb   = (1 + ipca_mensal) ** prazo_rec
    mora       = prazo_rec * 0.01
    fator_desc = (1 + taxa_desc_mensal) ** prazo_rec

    vj_bruto = np.where(fator_desc > 0, (vc * taxa_rec * (fc_receb + mora)) / fator_desc, 0.0)
    desc_aging = df["aging"].map(REMUNERATION_DISCOUNTS).fillna(0.50).astype(float)
    df["valor_justo"] = vj_bruto * (1 - desc_aging)

    return df


# ══════════════════════════════════════════════════════════════════════
# SUMÁRIO ESTATÍSTICO
# ══════════════════════════════════════════════════════════════════════

def compute_summary(df: pd.DataFrame) -> dict:
    def _s(c): return float(df[c].fillna(0).sum()) if c in df.columns else 0.0

    by_aging: dict = {}
    for label in AGING_LABELS:
        grp = df[df["aging"] == label] if "aging" in df.columns else pd.DataFrame()
        if grp.empty:
            continue
        by_aging[label] = {
            "count":           int(len(grp)),
            "valor_principal": float(grp["valor_principal_limpo"].fillna(0).sum()),
            "valor_corrigido": float(grp["valor_corrigido"].fillna(0).sum()),
            "valor_justo":     float(grp["valor_justo"].fillna(0).sum()),
        }

    by_empresa: dict = {}
    if "empresa" in df.columns:
        for emp, grp in df.groupby("empresa", sort=True):
            by_empresa[str(emp)] = {
                "count":           int(len(grp)),
                "valor_principal": float(grp["valor_principal_limpo"].fillna(0).sum()),
                "valor_corrigido": float(grp["valor_corrigido"].fillna(0).sum()),
                "valor_justo":     float(grp["valor_justo"].fillna(0).sum()),
            }

    return {
        "total_rows":               int(len(df)),
        "total_valor_principal":    _s("valor_principal_limpo"),
        "total_valor_liquido":      _s("valor_liquido"),
        "total_multa":              _s("multa"),
        "total_juros_moratorios":   _s("juros_moratorios"),
        "total_correcao_monetaria": _s("correcao_monetaria"),
        "total_valor_corrigido":    _s("valor_corrigido"),
        "total_valor_recuperavel":  _s("valor_recuperavel"),
        "total_valor_justo":        _s("valor_justo"),
        "by_aging":                 by_aging,
        "by_empresa":               by_empresa,
    }


# ══════════════════════════════════════════════════════════════════════
# EXPORTAÇÃO
# ══════════════════════════════════════════════════════════════════════

OUTPUT_COLS = [
    "empresa", "tipo", "nome_cliente", "documento", "contrato", "classe",
    "status_conta", "situacao",
    "valor_principal_limpo", "valor_nao_cedido", "valor_terceiro", "valor_cip",
    "data_vencimento", "data_base", "is_voltz",
    "dias_atraso", "aging", "aging_taxa",
    "valor_liquido", "multa", "juros_moratorios", "fator_correcao",
    "correcao_monetaria", "valor_corrigido",
    "juros_remuneratorios", "saldo_devedor_vencimento",
    "taxa_recuperacao", "prazo_recebimento", "valor_recuperavel", "valor_justo",
]

OUTPUT_HEADERS = [
    "Empresa", "Tipo", "Nome Cliente", "CPF/CNPJ", "Contrato", "Classe",
    "Status", "Situação",
    "Valor Principal", "Vlr Não Cedido", "Vlr Terceiros", "Vlr CIP",
    "Data Vencimento", "Data Base", "É Voltz",
    "Dias Atraso", "Aging", "Aging Taxa",
    "Valor Líquido", "Multa 2%", "Juros Moratórios", "Fator Correção",
    "Correção Monetária", "Valor Corrigido",
    "Juros Remun.", "SDV",
    "Taxa Recuperação", "Prazo Receb.", "Valor Recuperável", "Valor Justo",
]


def to_excel_bytes(df: pd.DataFrame, summary: dict) -> bytes:
    """Exporta resultado para Excel com aba de dados e aba de resumo."""
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Aba Dados
        out = df.reindex(columns=OUTPUT_COLS).copy()
        if "is_voltz" in out.columns:
            out["is_voltz"] = out["is_voltz"].map(lambda v: "Sim" if v else "Não")
        for col in ("data_vencimento", "data_base"):
            if col in out.columns:
                out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
        out.columns = OUTPUT_HEADERS[: len(out.columns)]
        out.to_excel(writer, sheet_name="Resultado", index=False)

        # Aba Resumo por Aging
        aging_rows = []
        for label in AGING_LABELS:
            d = summary["by_aging"].get(label, {})
            if d:
                aging_rows.append({
                    "Aging": label,
                    "Qtd": d["count"],
                    "Vlr Principal": d["valor_principal"],
                    "Vlr Corrigido": d["valor_corrigido"],
                    "Valor Justo":   d["valor_justo"],
                })
        if aging_rows:
            pd.DataFrame(aging_rows).to_excel(writer, sheet_name="Resumo Aging", index=False)

        # Aba Resumo por Empresa
        emp_rows = [
            {
                "Empresa": emp,
                "Qtd": d["count"],
                "Vlr Principal": d["valor_principal"],
                "Vlr Corrigido": d["valor_corrigido"],
                "Valor Justo":   d["valor_justo"],
            }
            for emp, d in sorted(summary["by_empresa"].items())
        ]
        if emp_rows:
            pd.DataFrame(emp_rows).to_excel(writer, sheet_name="Resumo Empresa", index=False)

    buf.seek(0)
    return buf.read()


def to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Exporta resultado para CSV semicolon, BOM UTF-8, formato pt-BR."""
    out = df.reindex(columns=OUTPUT_COLS).copy()
    if "is_voltz" in out.columns:
        out["is_voltz"] = out["is_voltz"].map(lambda v: "Sim" if v else "Não")
    for col in ("data_vencimento", "data_base"):
        if col in out.columns:
            out[col] = pd.to_datetime(out[col], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
    float_cols = [
        "valor_principal_limpo", "valor_nao_cedido", "valor_terceiro", "valor_cip",
        "valor_liquido", "multa", "juros_moratorios", "fator_correcao",
        "correcao_monetaria", "valor_corrigido",
        "juros_remuneratorios", "saldo_devedor_vencimento",
        "taxa_recuperacao", "valor_recuperavel", "valor_justo",
    ]
    for col in float_cols:
        if col in out.columns:
            out[col] = out[col].apply(lambda v: "" if pd.isna(v) else f"{float(v):.2f}".replace(".", ","))
    for col in out.select_dtypes(include="object").columns:
        out[col] = out[col].fillna("").astype(str).str.replace(";", ",", regex=False)
    out.columns = OUTPUT_HEADERS[: len(out.columns)]
    buf = io.StringIO()
    out.to_csv(buf, sep=";", index=False, lineterminator="\r\n")
    return ("\ufeff" + buf.getvalue()).encode("utf-8")
