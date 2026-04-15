"""
Página de Correção - FIDC Calculator
Cálculo de aging, correção monetária e valor justo
"""

import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime
import os
import time
from utils.calculador_aging import CalculadorAging
from utils.calculador_correcao import CalculadorCorrecao
from utils.calculador_valor_justo_distribuidoras import CalculadorValorJustoDistribuidoras, CalculadorValorJusto
from utils.visualizador_voltz import VisualizadorVoltz
from utils.visualizador_distribuidoras import VisualizadorDistribuidoras
from utils.auto_export_resultado import exportar_resultado_final_excel
from utils.exportacao_csv_brasil import salvar_csv_brasil
from utils.correcao_otimizada import (
    aplicar_correcao_monetaria_vetorizada,
    calcular_indice_diario_vetorizado,
    calcular_valor_justo_di_pre_vetorizado,
    otimizar_curva_di_pre,
)

# Importar classe de valor justo do app original
import requests
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from io import BytesIO

class CalculadorIndicesEconomicos:
    """
    Classe para cálculo de índices econômicos a partir de arquivo Excel
    """
    
    def __init__(self):
        self.df_indices = None
        self.ipca_12m_real = None
        self.data_base = None
    
    def carregar_indices_do_excel(self, arquivo_excel, aba_especifica=None):
        """
        Carrega dados de IGP-M/IPCA do arquivo Excel usando openpyxl para maior robustez.
        Suporta duas estruturas:
        1. Aba IGPM_IPCA: Coluna C (Ano), Coluna D (Mês) e Coluna F (Índice IGP-M)
        2. Aba IGPM: Coluna A (Mês/Ano), Coluna B (Índice)
        
        Parâmetros:
        - arquivo_excel: arquivo Excel para carregar
        - aba_especifica: nome específico da aba (para VOLTZ usar "IGPM")
        """
        try:
            import pandas as pd
            import streamlit as st
            import numpy as np
            from datetime import datetime
            from openpyxl import load_workbook
            from io import BytesIO
            
            print("🔄 Carregando índices do arquivo Excel com openpyxl...")
            
            # ===== PREPARAR ARQUIVO PARA OPENPYXL =====
            if hasattr(arquivo_excel, 'read'):
                # Se é um objeto file, ler bytes
                if hasattr(arquivo_excel, 'seek'):
                    arquivo_excel.seek(0)  # Voltar para o início
                arquivo_bytes = BytesIO(arquivo_excel.read())
            else:
                arquivo_bytes = arquivo_excel
            
            # ===== CARREGAR WORKBOOK COM OPENPYXL =====
            try:
                workbook = load_workbook(arquivo_bytes, data_only=True, read_only=True)
                print(f"📊 Workbook carregado. Abas disponíveis: {workbook.sheetnames}")
            except Exception as e:
                print(f"❌ Erro ao carregar workbook: {e}")
                st.error(f"Erro ao abrir arquivo Excel: {e}")
                return None
            
            # ===== DETERMINAR QUAL ABA USAR =====
            if aba_especifica:
                if aba_especifica in workbook.sheetnames:
                    sheet_name = aba_especifica
                    print(f"✅ Aba específica '{aba_especifica}' encontrada")
                else:
                    print(f"⚠️ Aba '{aba_especifica}' não encontrada. Abas disponíveis: {workbook.sheetnames}")
                    # Tentar primeira aba como fallback
                    sheet_name = workbook.sheetnames[0]
                    print(f"📊 Usando primeira aba como fallback: '{sheet_name}'")
            else:
                # Usar primeira aba
                sheet_name = workbook.sheetnames[0]
                print(f"📊 Usando primeira aba: '{sheet_name}'")
            
            # ===== CARREGAR DADOS DA ABA SELECIONADA =====
            try:
                worksheet = workbook[sheet_name]
                print(f"📊 Aba '{sheet_name}' carregada com {worksheet.max_row} linhas e {worksheet.max_column} colunas")
                
                # Converter worksheet para lista de listas para processamento
                dados_brutos = []
                for row in worksheet.iter_rows(values_only=True):
                    dados_brutos.append(list(row))
                
                # Criar DataFrame a partir dos dados brutos
                df_completo = pd.DataFrame(dados_brutos)
                
                # Remover linhas completamente vazias
                df_completo = df_completo.dropna(how='all').reset_index(drop=True)
                
                print(f"📊 Dados extraídos: {df_completo.shape[0]} linhas e {df_completo.shape[1]} colunas (após limpeza)")
                
            except Exception as e:
                print(f"❌ Erro ao ler aba '{sheet_name}': {e}")
                st.error(f"Erro ao ler aba '{sheet_name}': {e}")
                return None
            finally:
                workbook.close()
            
            # ===== DETECTAR E PROCESSAR ESTRUTURA DOS DADOS =====
            if aba_especifica == "IGPM":
                print("📊 Processando estrutura IGPM - Colunas A (data) e B (índice)")
                
                if df_completo.shape[1] < 2:
                    raise ValueError(f"Aba IGPM tem apenas {df_completo.shape[1]} colunas, mas precisa de pelo menos 2 (A-B)")
                
                # Extrair colunas A (0) e B (1) - Data e Índice
                df = df_completo.iloc[:, [0, 1]].copy()
                df.columns = ['data_excel', 'indice']
                
                # Remover linhas com dados nulos
                df = df.dropna(subset=['data_excel']).copy()
                
                print(f"📊 Dados extraídos: {len(df)} linhas com data_excel e índice")
                
                # ===== PROCESSAR DIFERENTES FORMATOS DE DATA =====
                df['ano'] = None
                df['mes'] = None
                
                for idx, row in df.iterrows():
                    data_valor = row['data_excel']
                    
                    try:
                        # Caso 1: Data do Excel (datetime ou serial)
                        if isinstance(data_valor, datetime):
                            df.at[idx, 'ano'] = data_valor.year
                            df.at[idx, 'mes'] = data_valor.month
                            print(f"✅ Linha {idx}: Data Excel datetime - {data_valor.year}/{data_valor.month}")
                            
                        elif isinstance(data_valor, (int, float)):
                            # Converter serial do Excel para datetime
                            from datetime import timedelta
                            # Excel epoch: 1900-01-01 (mas com bug do ano 1900)
                            excel_epoch = datetime(1899, 12, 30)  # Correção do bug do Excel
                            data_convertida = excel_epoch + timedelta(days=int(data_valor))
                            df.at[idx, 'ano'] = data_convertida.year
                            df.at[idx, 'mes'] = data_convertida.month
                            print(f"✅ Linha {idx}: Serial Excel {data_valor} → {data_convertida.year}/{data_convertida.month}")
                            
                        elif isinstance(data_valor, str):
                            # Caso 2: Formato texto "mês/ano"
                            if '/' in str(data_valor):
                                meses_dict = {
                                    'janeiro': 1, 'fevereiro': 2, 'março': 3, 'abril': 4, 'maio': 5, 'junho': 6,
                                    'julho': 7, 'agosto': 8, 'setembro': 9, 'outubro': 10, 'novembro': 11, 'dezembro': 12
                                }
                                
                                partes = str(data_valor).split('/')
                                if len(partes) >= 2:
                                    mes_nome = partes[0].strip().lower()
                                    ano_str = partes[1].strip()
                                    
                                    if mes_nome in meses_dict:
                                        df.at[idx, 'mes'] = meses_dict[mes_nome]
                                        df.at[idx, 'ano'] = int(ano_str)
                                        print(f"✅ Linha {idx}: Texto '{data_valor}' → {int(ano_str)}/{meses_dict[mes_nome]}")
                                    else:
                                        print(f"⚠️ Linha {idx}: Mês '{mes_nome}' não reconhecido")
                                else:
                                    print(f"⚠️ Linha {idx}: Formato inválido '{data_valor}'")
                            else:
                                # Tentar converter string para datetime
                                try:
                                    data_convertida = pd.to_datetime(str(data_valor), errors='raise')
                                    df.at[idx, 'ano'] = data_convertida.year
                                    df.at[idx, 'mes'] = data_convertida.month
                                    print(f"✅ Linha {idx}: String '{data_valor}' → {data_convertida.year}/{data_convertida.month}")
                                except:
                                    print(f"⚠️ Linha {idx}: Não foi possível converter '{data_valor}'")
                        else:
                            print(f"⚠️ Linha {idx}: Tipo não suportado {type(data_valor)}: {data_valor}")
                            
                    except Exception as e:
                        print(f"❌ Linha {idx}: Erro ao processar '{data_valor}': {e}")
                        continue
                
                # Remover linhas onde não conseguimos extrair ano e mês
                df = df.dropna(subset=['ano', 'mes']).copy()
                df['ano'] = df['ano'].astype(int)
                df['mes'] = df['mes'].astype(int)
                
                print(f"📊 Processamento IGPM: {len(df)} registros com data válida")
                
            else:
                print("📊 Processando estrutura IGPM_IPCA - Colunas C (ano), D (mês) e F (índice)")
                
                # Verificar se tem pelo menos 6 colunas (A até F)
                if df_completo.shape[1] < 6:
                    raise ValueError(f"Estrutura IGPM_IPCA tem apenas {df_completo.shape[1]} colunas, mas precisa de pelo menos 6 (A-F)")
                
                # Extrair colunas C (2), D (3) e F (5) - Ano, Mês e Índice IGP-M
                df = df_completo.iloc[:, [2, 3, 5]].copy()
                df.columns = ['ano', 'mes', 'indice']
            
            # ===== PADRONIZAR COLUNAS PARA ['ano', 'mes', 'indice'] =====
            if 'data_excel' in df.columns:
                # Para estrutura IGPM, já processamos e temos ano, mes, indice
                df = df[['ano', 'mes', 'indice']].copy()
            elif 'mes_ano' in df.columns:
                # Fallback para formato antigo (se ainda existir)
                df = df[['ano', 'mes', 'indice']].copy()
            
            print(f"📊 Dados padronizados - ano, mês e índice extraídos")
            
            # ===== LIMPEZA E VALIDAÇÃO DOS DADOS =====
            # Limpar dados nulos
            df_original_len = len(df)
            df = df.dropna().copy()
            print(f"📊 Removidas {df_original_len - len(df)} linhas com valores nulos")
            
            # Processar colunas de ano e mês
            print("🔄 Processando e validando colunas de ano e mês...")
            
            # Converter ano e mês para inteiros com tratamento robusto
            try:
                df['ano'] = pd.to_numeric(df['ano'], errors='coerce').astype('Int64')
                df['mes'] = pd.to_numeric(df['mes'], errors='coerce').astype('Int64')
                
                print(f"📊 Conversão bem-sucedida - Anos: {sorted(df['ano'].dropna().unique())[:5]}...")
                print(f"📊 Meses únicos: {sorted(df['mes'].dropna().unique())}")
                
            except Exception as conv_error:
                print(f"⚠️ Erro na conversão padrão: {conv_error}")
                # Tentativa de conversão manual mais robusta
                df['ano'] = df['ano'].astype(str).str.extract(r'(\d{4})')[0].astype(float).astype('Int64')
                df['mes'] = df['mes'].astype(str).str.extract(r'(\d{1,2})')[0].astype(float).astype('Int64')
                print("✅ Conversão manual aplicada com sucesso")
            
            # Remover linhas com ano/mês inválidos
            df_limpo = df.dropna(subset=['ano', 'mes']).copy()
            
            # Validar valores de ano e mês
            mask_ano_valido = (df_limpo['ano'] >= 1990) & (df_limpo['ano'] <= 2030)
            mask_mes_valido = (df_limpo['mes'] >= 1) & (df_limpo['mes'] <= 12)
            
            df_filtrado = df_limpo[mask_ano_valido & mask_mes_valido].copy()
            
            if len(df_filtrado) == 0:
                raise ValueError("Nenhuma data válida encontrada após validação")
            
            print(f"📊 {len(df_filtrado)} registros com ano/mês válidos")
            
            # ===== CRIAR COLUNA DE DATA =====
            try:
                df_filtrado['data'] = pd.to_datetime(
                    df_filtrado['ano'].astype(str) + '-' + df_filtrado['mes'].astype(str).str.zfill(2) + '-01',
                    format='%Y-%m-%d'
                )
                print(f"📊 {len(df_filtrado)} datas criadas com sucesso")
            except Exception as e:
                print(f"❌ Erro ao criar datas: {e}")
                raise ValueError("Erro ao combinar ano e mês em datas válidas")
            
            # ===== PROCESSAR COLUNA DE ÍNDICE =====
            print("🔄 Processando e validando coluna de índice...")
            
            # Converter índice para float, tratando vírgulas decimais brasileiras
            if df_filtrado['indice'].dtype == 'object':
                df_filtrado['indice'] = df_filtrado['indice'].astype(str).str.replace(',', '.')
            
            # Converter para numérico
            df_filtrado['indice'] = pd.to_numeric(df_filtrado['indice'], errors='coerce')
            
            # Remover linhas com índices inválidos
            df_final = df_filtrado.dropna(subset=['indice']).copy()
            
            if len(df_final) == 0:
                raise ValueError("Nenhum índice válido encontrado após conversão numérica")
            
            print(f"📊 Índices convertidos: {len(df_filtrado)} → {len(df_final)} registros válidos")
            
            # ===== FINALIZAR E VALIDAR =====
            # Manter apenas colunas necessárias
            df_final = df_final[['data', 'indice']].copy()
            
            # Ordenar por data
            df_final = df_final.sort_values('data').reset_index(drop=True)
            
            # Validações finais
            periodo_min = df_final['data'].min()
            periodo_max = df_final['data'].max()
            indice_min = df_final['indice'].min()
            indice_max = df_final['indice'].max()
            
            # Salvar resultado
            self.df_indices = df_final
            
            print(f"✅ {len(df_final)} registros carregados com OPENPYXL!")
            print(f"📊 Período: {periodo_min.strftime('%Y-%m')} a {periodo_max.strftime('%Y-%m')}")
            print(f"📊 Índice: {indice_min:.2f} a {indice_max:.2f}")
            
            return self.df_indices
            
        except Exception as e:
            error_msg = f"❌ Erro ao carregar índices com openpyxl: {str(e)}"
            print(error_msg)
            if 'st' in locals():
                st.error(f"Erro ao processar arquivo Excel: {str(e)}")
            return None
    
    
    def buscar_indice_para_data(self, data_busca):
        """
        Busca o índice para uma data específica
        """
        if self.df_indices is None or self.df_indices.empty:
            print("⚠️ Dados de índices não carregados")
            return None
        
        try:
            # Converter data para período mensal para busca
            periodo_busca = pd.Period(data_busca, freq='M')
            
            # Adicionar coluna de período ao dataframe se não existir
            if 'periodo' not in self.df_indices.columns:
                self.df_indices['periodo'] = self.df_indices['data'].dt.to_period('M')
            
            # Buscar índice exato para o período
            resultado = self.df_indices[self.df_indices['periodo'] == periodo_busca]
            
            if not resultado.empty:
                return resultado.iloc[0]['indice']
            else:
                # Se não encontrar exato, buscar o mais próximo anterior
                indices_anteriores = self.df_indices[self.df_indices['periodo'] <= periodo_busca]
                if not indices_anteriores.empty:
                    return indices_anteriores.iloc[-1]['indice']
                else:
                    print(f"⚠️ Não foi possível encontrar índice para {data_busca.strftime('%Y-%m')}")
                    return None
                    
        except Exception as e:
            print(f"❌ Erro ao buscar índice para {data_busca}: {e}")
            return None
    
    def calcular_fator_correcao(self, data_vencimento, data_base):
        """
        Calcula o fator de correção entre duas datas usando os índices carregados
        """
        try:
            indice_vencimento = self.buscar_indice_para_data(data_vencimento)
            indice_base = self.buscar_indice_para_data(data_base)
            
            if indice_vencimento is None or indice_base is None:
                print(f"⚠️ Não foi possível encontrar índices para calcular correção")
                return 1.0  # Retorna fator neutro
            
            # Fator de correção = índice_base / índice_vencimento
            fator = indice_base / indice_vencimento
            
            return fator
            
        except Exception as e:
            print(f"❌ Erro ao calcular fator de correção: {e}")
            return 1.0

class CalculadorValorJusto:
    """
    Classe para cálculo do valor justo usando índices do Excel
    """
    
    def __init__(self):
        self.calculador_indices = CalculadorIndicesEconomicos()
        
    def carregar_dados_indices(self, arquivo_excel):
        """
        Carrega dados de índices do arquivo Excel.
        Para VOLTZ: carrega aba IGPM e salva em df_indices_igpm
        Para outras distribuidoras: carrega aba IGPM_IPCA e salva em df_indices_economicos
        """
        # Detectar se é VOLTZ baseado nos arquivos carregados
        eh_voltz = False
        
        if 'df_dados_principais' in st.session_state:
            dados_principais = st.session_state.df_dados_principais
            if not dados_principais.empty and hasattr(dados_principais, 'attrs'):
                nome_arquivo = dados_principais.attrs.get('nome_arquivo', '')
                if 'VOLTZ' in nome_arquivo.upper():
                    eh_voltz = True
        
        if not eh_voltz and 'distribuidora_detectada' in st.session_state:
            if st.session_state.distribuidora_detectada == 'VOLTZ':
                eh_voltz = True
        
        if eh_voltz:
            # VOLTZ: Carregar apenas aba IGPM
            st.info("🔍 **VOLTZ detectado** - Carregando dados da aba 'IGPM'")
            df_igpm = self.calculador_indices.carregar_indices_do_excel(arquivo_excel, "IGPM")
            if df_igpm is not None and not df_igpm.empty:
                # Salvar dados IGP-M específicos para VOLTZ
                st.session_state.df_indices_igpm = df_igpm
                # Também salvar em df_indices_economicos para compatibilidade
                st.session_state.df_indices_economicos = df_igpm
                st.success("⚡ **VOLTZ**: Dados IGP-M da aba 'IGPM' carregados com sucesso!")
                return df_igpm
            else:
                st.error("❌ Erro ao carregar dados da aba 'IGPM' para VOLTZ")
                return None
        else:
            # DISTRIBUIDORAS: Carregar aba IGPM_IPCA (padrão)
            st.info("🏢 **Distribuidora Geral** - Carregando dados da primeira aba (IGPM_IPCA)")
            df_ipca = self.calculador_indices.carregar_indices_do_excel(arquivo_excel, None)  # Primeira aba
            if df_ipca is not None and not df_ipca.empty:
                # Salvar dados IPCA para distribuidoras gerais
                st.session_state.df_indices_economicos = df_ipca
                st.success("🏢 **Distribuidoras**: Dados de índices carregados com sucesso!")
                return df_ipca
            else:
                st.error("❌ Erro ao carregar dados de índices para distribuidoras")
                return None
    
    def obter_indice_para_data(self, data_busca):
        """
        Obtém índice para uma data específica
        """
        return self.calculador_indices.buscar_indice_para_data(data_busca)
    
    def calcular_fator_correcao_indices(self, data_vencimento, data_base):
        """
        Calcula fator de correção usando os índices carregados
        """
        return self.calculador_indices.calcular_fator_correcao(data_vencimento, data_base)
    
    def calcular_valor_justo_com_di_pre(self, df_corrigido, df_di_pre, coluna_valor_corrigido='valor_corrigido', data_base=None):
        """
        Calcula valor justo aplicando taxas DI-PRE com progressão exponencial sobre valor corrigido
        """
        try:
            return calcular_valor_justo_di_pre_vetorizado(
                df_corrigido=df_corrigido,
                df_di_pre=df_di_pre,
                coluna_valor_corrigido=coluna_valor_corrigido,
                data_base=data_base,
            )
        except KeyError:
            st.error("❌ Taxa de recuperação não encontrada nos dados")
            return
    
    def obter_estatisticas_di_pre(self, df_di_pre):
        """
        Retorna estatísticas do DI-PRE calculado incluindo informações sobre progressão exponencial
        """
        if df_di_pre.empty:
            return None
        
        # Calcular estatísticas básicas das taxas DI-PRE (coluna '252')
        taxa_media = df_di_pre['252'].mean()
        taxa_min = df_di_pre['252'].min()
        taxa_max = df_di_pre['252'].max()
        
        return {
            'taxa_media_percentual': taxa_media,
            'taxa_min_percentual': taxa_min,
            'taxa_max_percentual': taxa_max,
            'total_registros_di_pre': len(df_di_pre),
            'meses_disponiveis': sorted(df_di_pre['meses_futuros'].unique().tolist()),
            'formula_exponencial': 'An = (1 + taxa_di_pre)^(prazo_recebimento/12)',
            'descricao_prazo': 'prazo_recebimento em meses'
        }

def show():
    """Página de Correção Monetária e Valor Justo"""
    st.header("💰 CORREÇÃO MONETÁRIA e VALOR JUSTO")
    
    # Inicializar flags de controle
    if 'calculo_solicitado' not in st.session_state:
        st.session_state.calculo_solicitado = False
    if 'taxa_recuperacao_carregada' not in st.session_state:
        st.session_state.taxa_recuperacao_carregada = False
    if 'cdi_carregado' not in st.session_state:
        st.session_state.cdi_carregado = False
    if 'indices_carregados' not in st.session_state:
        st.session_state.indices_carregados = False
    
    # Verificar se temos dados padronizados
    if 'df_padronizado' not in st.session_state or st.session_state.df_padronizado.empty:
        st.warning("⚠️ Realize o mapeamento de campos antes de calcular a correção monetária.")
        st.info("💡 Vá para a página de **Mapeamento** e complete o processo de mapeamento primeiro.")
        return
    
    # Verificar se os parâmetros estão inicializados
    if 'params' not in st.session_state:
        from utils.parametros_correcao import ParametrosCorrecao
        st.session_state.params = ParametrosCorrecao()
    
    df_padronizado = st.session_state.df_padronizado
    calc_aging = CalculadorAging(st.session_state.params)
    calc_correcao = CalculadorCorrecao(st.session_state.params)
    
    # ETAPA 0: CARREGAMENTO DOS ÍNDICES IGP-M/IPCA (OBRIGATÓRIO)
    st.subheader("📊 0️⃣ Carregar Índices IGP-M/IPCA")
    
    # Verificar se temos dados de índices carregados (qualquer uma das abas)
    tem_indices = (
        ('df_indices_economicos' in st.session_state and not st.session_state.df_indices_economicos.empty) or
        ('df_indices_igpm' in st.session_state and not st.session_state.df_indices_igpm.empty)
    )
    
    if not tem_indices:
        st.warning("⚠️ **PASSO 0:** Faça o upload do arquivo de índices IGP-M/IPCA para continuar.")
        
        with st.expander("📤 Upload dos Índices IGP-M/IPCA", expanded=True):
            st.info("""
            **📋 Instruções:** 
            
            Faça o upload do arquivo Excel com os índices econômicos. O sistema suporta duas estruturas:
            
            **Estrutura 1 (IGPM_IPCA):**
            - **Coluna C**: Ano (ex: 2022)
            - **Coluna D**: Mês (ex: 1, 2, 3...)  
            - **Coluna F**: Valores dos índices IGP-M ou IPCA
            
            **Estrutura 2 (IGPM - para VOLTZ):**
            - **Coluna A**: Mês/Ano (ex: agosto/1994, setembro/1994)
            - **Coluna B**: Índice (ex: 100.000, 101.751)
            
            O sistema detecta automaticamente qual estrutura usar baseado na distribuidora.
            
            **Exemplo de estrutura IGPM_IPCA:**
            ```
            2021    12    543.21
            2022    01    548.65  
            2022    02    552.34
            ```
            
            **Exemplo de estrutura IGPM:**
            ```
            agosto/1994      100.000
            setembro/1994    101.751  
            outubro/1994     103.602
            ```
            """)
            
            # Upload do arquivo
            uploaded_file_indices = st.file_uploader(
                "📤 Selecione o arquivo de Índices IGP-M/IPCA (ambas as abas serão carregadas)",
                type=['xlsx', 'xls'],
                help="Arquivo Excel com índices econômicos. Sistema carregará automaticamente abas IGPM e IGPM_IPCA",
                key="upload_indices_modulo4"
            )
            
            if uploaded_file_indices is not None:
                try:
                    with st.spinner("🔄 Processando arquivo de índices (carregando ambas as abas)..."):
                        # Criar instância do calculador
                        calc_valor_justo = CalculadorValorJusto()
                        
                        # ===== CARREGAR AMBAS AS ABAS SEPARADAMENTE =====
                        st.info("📊 **Carregando aba IGPM** (para VOLTZ)...")
                        df_igpm = calc_valor_justo.calculador_indices.carregar_indices_do_excel(uploaded_file_indices, "IGPM")
                        
                        st.info("📊 **Carregando aba IGMP_IPCA** (para outras distribuidoras)...")  
                        df_ipca = calc_valor_justo.calculador_indices.carregar_indices_do_excel(uploaded_file_indices, None)  # Primeira aba
                        
                        # ===== VALIDAR E SALVAR DADOS =====
                        dados_carregados = 0
                        
                        if df_igpm is not None and not df_igpm.empty:
                            st.session_state.df_indices_igpm = df_igpm
                            dados_carregados += 1
                            st.success(f"✅ **Aba IGPM carregada:** {len(df_igpm):,} registros (para VOLTZ)")
                        else:
                            st.warning("⚠️ **Aba IGPM não encontrada** ou vazia")
                        
                        if df_ipca is not None and not df_ipca.empty:
                            st.session_state.df_indices_economicos = df_ipca
                            dados_carregados += 1
                            st.success(f"✅ **Aba IGPM_IPCA carregada:** {len(df_ipca):,} registros (para distribuidoras)")
                        else:
                            st.warning("⚠️ **Aba IGPM_IPCA não encontrada** ou vazia")
                        
                        # ===== VERIFICAR SE PELO MENOS UMA ABA FOI CARREGADA =====
                        if dados_carregados > 0:
                            st.session_state.calculador_valor_justo = calc_valor_justo
                            st.session_state.indices_carregados = True
                            
                            st.success(f"🎯 **{dados_carregados}/2 abas carregadas com sucesso!**")
                            
                            # ===== MOSTRAR ESTATÍSTICAS DE AMBAS AS ABAS =====
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                if 'df_indices_igpm' in st.session_state:
                                    df_stats = st.session_state.df_indices_igpm
                                    st.write("**📊 IGPM (VOLTZ):**")
                                    st.metric("Registros", f"{len(df_stats):,}")
                                    if not df_stats.empty:
                                        periodo_min = df_stats['data'].min().strftime('%Y-%m')
                                        periodo_max = df_stats['data'].max().strftime('%Y-%m')
                                        st.metric("Período", f"{periodo_min} a {periodo_max}")
                                        indice_atual = df_stats['indice'].iloc[-1]
                                        st.metric("Último Índice", f"{indice_atual:.2f}")
                                else:
                                    st.write("**❌ IGPM: Não carregado**")
                            
                            with col2:
                                if 'df_indices_economicos' in st.session_state:
                                    df_stats = st.session_state.df_indices_economicos
                                    st.write("**📊 IGPM_IPCA (Distribuidoras):**")
                                    st.metric("Registros", f"{len(df_stats):,}")
                                    if not df_stats.empty:
                                        periodo_min = df_stats['data'].min().strftime('%Y-%m')
                                        periodo_max = df_stats['data'].max().strftime('%Y-%m')
                                        st.metric("Período", f"{periodo_min} a {periodo_max}")
                                        indice_atual = df_stats['indice'].iloc[-1]
                                        st.metric("Último Índice", f"{indice_atual:.2f}")
                                else:
                                    st.write("**❌ IGPM_IPCA: Não carregado**")
                            
                            # ===== MOSTRAR PREVIEW DAS DUAS ABAS =====
                            st.write("**📋 Preview dos Dados Carregados:**")
                            
                            tab1, tab2 = st.tabs(["📊 IGPM (VOLTZ)", "📊 IGPM_IPCA (Distribuidoras)"])
                            
                            with tab1:
                                if 'df_indices_igpm' in st.session_state:
                                    st.dataframe(st.session_state.df_indices_igpm.head(10), use_container_width=True)
                                else:
                                    st.info("⚠️ Aba IGPM não foi carregada")
                            
                            with tab2:
                                if 'df_indices_economicos' in st.session_state:
                                    st.dataframe(st.session_state.df_indices_economicos.head(10), use_container_width=True)
                                else:
                                    st.info("⚠️ Aba IGPM_IPCA não foi carregada")
                            
                            st.rerun()
                        else:
                            st.error("❌ Nenhuma aba válida foi encontrada no arquivo Excel.")
                            st.error("🔧 Verifique se o arquivo possui abas 'IGPM' e/ou dados na primeira aba.")
                            
                except Exception as e:
                    st.error(f"❌ Erro ao processar arquivo: {str(e)}")
                    st.error("🔧 Verifique se o arquivo possui as abas corretas e dados válidos.")

        # Se não tem índices, parar aqui
        return
    else:
        st.success("✅ **Índices IGP-M/IPCA carregados**")
        
        
        # Mostrar informações das abas carregadas
        abas_carregadas = []
        if 'df_indices_igpm' in st.session_state and not st.session_state.df_indices_igpm.empty:
            registros_igpm = len(st.session_state.df_indices_igpm)
            periodo_min_igpm = st.session_state.df_indices_igpm['data'].min().strftime('%Y-%m')
            periodo_max_igpm = st.session_state.df_indices_igpm['data'].max().strftime('%Y-%m')
            abas_carregadas.append(f"IGPM: {registros_igpm:,} registros ({periodo_min_igpm} a {periodo_max_igpm})")
        
        if 'df_indices_economicos' in st.session_state and not st.session_state.df_indices_economicos.empty:
            registros_ipca = len(st.session_state.df_indices_economicos)
            periodo_min_ipca = st.session_state.df_indices_economicos['data'].min().strftime('%Y-%m')
            periodo_max_ipca = st.session_state.df_indices_economicos['data'].max().strftime('%Y-%m')
            abas_carregadas.append(f"IGPM_IPCA: {registros_ipca:,} registros ({periodo_min_ipca} a {periodo_max_ipca})")
        
        for info in abas_carregadas:
            st.info(f"📊 {info}")
        
        if st.button("🔄 Recarregar Índices", key="recarregar_indices"):
            st.session_state.indices_carregados = False
            st.session_state.taxa_recuperacao_carregada = False
            st.session_state.cdi_carregado = False
            st.session_state.calculo_solicitado = False
            # Limpar ambos os DataFrames
            if 'df_indices_economicos' in st.session_state:
                del st.session_state.df_indices_economicos
            if 'df_indices_igpm' in st.session_state:
                del st.session_state.df_indices_igpm
            if 'calculador_valor_justo' in st.session_state:
                del st.session_state.calculador_valor_justo
            st.rerun()

        # Mostrar preview das abas carregadas
        with st.expander("📊 Preview dos Índices Carregados", expanded=False):
            if 'df_indices_igpm' in st.session_state and 'df_indices_economicos' in st.session_state:
                # Mostrar ambas as abas
                tab1, tab2 = st.tabs(["📊 IGPM (VOLTZ)", "📊 IGPM_IPCA (Distribuidoras)"])
                
                with tab1:
                    st.dataframe(st.session_state.df_indices_igpm, use_container_width=True)
                
                with tab2:
                    st.dataframe(st.session_state.df_indices_economicos, use_container_width=True)
            
            elif 'df_indices_igpm' in st.session_state:
                st.write("**📊 IGPM (VOLTZ):**")
                st.dataframe(st.session_state.df_indices_igpm, use_container_width=True)
            
            elif 'df_indices_economicos' in st.session_state:
                st.write("**📊 IGPM_IPCA (Distribuidoras):**")
                st.dataframe(st.session_state.df_indices_economicos, use_container_width=True)

    st.markdown("---")
    
    # ETAPA 1: CARREGAMENTO DA TAXA DE RECUPERAÇÃO (OBRIGATÓRIO)
    st.subheader("📈 1️⃣ Configurar Taxa de Recuperação")
    
    # Verificar se temos dados de taxa de recuperação
    tem_taxa_recuperacao = 'df_taxa_recuperacao' in st.session_state and not st.session_state.df_taxa_recuperacao.empty
    
    if not tem_taxa_recuperacao:
        st.warning("⚠️ **PASSO 1:** Faça o upload do arquivo de taxa de recuperação para continuar.")
        
        with st.expander("📤 Upload da Taxa de Recuperação", expanded=True):
            st.info("""
            **📋 Instruções:** 
            
            Faça o upload do arquivo Excel com as taxas de recuperação. O arquivo deve conter:
            - Uma aba chamada "Input" 
            - Estrutura com empresas marcadas com "x" 
            - Tipos: Privada, Público, Hospital
            - Aging: A vencer, Primeiro ano, Segundo ano, Terceiro ano, Demais anos
            - Taxas e prazos de recebimento
            """)
            
            # Upload do arquivo
            uploaded_file_taxa = st.file_uploader(
                "📤 Selecione o arquivo de Taxa de Recuperação",
                type=['xlsx', 'xls'],
                help="Arquivo Excel com as taxas de recuperação por empresa, tipo e aging",
                key="upload_taxa_modulo4"
            )
            
            if uploaded_file_taxa is not None:
                try:
                    with st.spinner("🔄 Processando arquivo de taxa de recuperação..."):
                        # [Código de processamento mantido igual]
                        df_taxa_upload = pd.read_excel(uploaded_file_taxa, sheet_name="Input", header=None)
                        
                        tipos = ["Privado", "Público", "Hospital"]
                        aging_labels = ["A vencer", "Primeiro ano", "Segundo ano", "Terceiro ano", "Quarto ano", "Quinto ano", "Demais anos"]

                        empresa = None
                        dados_taxa = []
                        
                        for i in range(len(df_taxa_upload)):
                            row = df_taxa_upload.iloc[i]

                            for j in range(len(row) - 1):
                                if str(row[j]).strip().lower() == "x":
                                    empresa = str(row[j + 1]).strip()

                            if not empresa:
                                continue

                            for offset, tipo in zip([1, 5, 9], tipos):
                                try:
                                    aging = str(row[offset]).strip()
                                    taxa = row[offset + 1]
                                    prazo = row[offset + 2]

                                    if aging in aging_labels and pd.notna(taxa) and pd.notna(prazo):
                                        dados_taxa.append({
                                            "Empresa": empresa,
                                            "Tipo": tipo,
                                            "Aging": aging,
                                            "Taxa de recuperação": float(str(taxa).replace(",", ".")),
                                            "Prazo de recebimento": int(prazo)
                                        })
                                except (IndexError, ValueError):
                                    continue
                        
                        if dados_taxa:
                            df_taxa_recuperacao_nova = pd.DataFrame(dados_taxa)
                            st.session_state.df_taxa_recuperacao = df_taxa_recuperacao_nova
                            st.session_state.taxa_recuperacao_carregada = True
                            
                            # Resetar flags subsequentes
                            st.session_state.cdi_carregado = False
                            st.session_state.calculo_solicitado = False
                            if 'df_final' in st.session_state:
                                del st.session_state.df_final
                            if 'df_com_aging' in st.session_state:
                                del st.session_state.df_com_aging
                            
                            st.rerun()
                        else:
                            st.error("❌ Nenhum dado válido encontrado no arquivo. Verifique a estrutura do arquivo.")
                            
                except Exception as e:
                    st.error(f"❌ Erro ao processar arquivo: {str(e)}")
                    st.error("Verifique se o arquivo possui uma aba 'Input' e se a estrutura está correta.")

        # Se não tem taxa, parar aqui
        return
    else:
        st.success("✅ **Taxa de recuperação carregada**")
        empresas = st.session_state.df_taxa_recuperacao['Empresa'].nunique()
        registros = len(st.session_state.df_taxa_recuperacao)
        st.info(f"🏢 {empresas} empresa(s) configurada(s) com {registros} registro(s) de taxa")
        
        if st.button("🔄 Recarregar Taxa de Recuperação", key="recarregar_taxa_recuperacao"):
            st.session_state.taxa_recuperacao_carregada = False
            st.session_state.cdi_carregado = False
            st.session_state.calculo_solicitado = False
            if 'df_taxa_recuperacao' in st.session_state:
                del st.session_state.df_taxa_recuperacao
            st.rerun()

        # Mostrar o head dos dados num expander
        with st.expander("📊 Preview dos Dados de Taxa de Recuperação", expanded=False):
            if 'df_taxa_recuperacao' in st.session_state:
                st.dataframe(st.session_state.df_taxa_recuperacao, use_container_width=True)

    st.markdown("---")
    
    # Detectar se é VOLTZ para determinar se precisa de CDI
    nome_arquivo_detectado = "Distribuidora"  # Default
    if 'df_carregado' in st.session_state and st.session_state.df_carregado:
        primeiro_arquivo = list(st.session_state.df_carregado.keys())[0]
        nome_arquivo_detectado = primeiro_arquivo
    
    # Verificar se é VOLTZ
    from utils.calculador_voltz import CalculadorVoltz
    calculador_voltz = CalculadorVoltz(st.session_state.params)
    eh_voltz = calculador_voltz.identificar_voltz(nome_arquivo_detectado)
    
    # ETAPA 2: CARREGAMENTO DO ARQUIVO CDI (OBRIGATÓRIO PARA TODAS AS DISTRIBUIDORAS)
    st.subheader("📈 2️⃣ Carregar Dados CDI/DI-PRE")
    
    if eh_voltz:
        st.info("⚡ **VOLTZ detectada:** CDI/DI-PRE necessário para cálculo do valor justo (desconto a valor presente)")
    
    # Verificar se temos dados CDI carregados
    tem_cdi = 'df_di_pre' in st.session_state and not st.session_state.df_di_pre.empty
    
    if not tem_cdi:
        st.warning("⚠️ **PASSO 2:** Faça o upload do arquivo CDI/DI-PRE para continuar.")
        
        with st.expander("📤 Upload do Arquivo CDI/DI-PRE", expanded=True):
            st.info("""
            **📋 Instruções:** 
            
            Faça o upload do arquivo Excel (.xls ou .xlsx) com os dados de CDI/DI-PRE da BMF.
            Este arquivo contém as taxas de juros utilizadas para correção monetária.
            
            **Formato esperado:** Arquivo HTML/Excel da BMF com dados de DI x pré
            """)
            st.markdown(
                "🔗 Fonte oficial DI-PRE (B3): https://www2.bmf.com.br/pages/portal/bmfbovespa/boletim1/txref1.asp"
            )
            
            # Upload do arquivo CDI
            uploaded_file_cdi = st.file_uploader(
                "📤 Selecione o arquivo CDI/DI-PRE",
                type=['xlsx', 'xls'],
                help="Arquivo Excel com dados de CDI/DI-PRE da BMF",
                key="upload_cdi_modulo4"
            )
            
            if uploaded_file_cdi is not None:
                try:
                    with st.spinner("🔄 Processando arquivo CDI/DI-PRE..."):
                        # Importar e usar o processador CDI
                        from utils.processador_di_pre import ProcessadorDIPre
                        
                        processador = ProcessadorDIPre()
                        df_cdi_processado = processador.processar_arquivo_bmf(uploaded_file_cdi)
                        
                        if df_cdi_processado is not None and not df_cdi_processado.empty:
                            st.session_state.df_di_pre = df_cdi_processado
                            st.session_state.cdi_carregado = True
                            
                            # Resetar flag de cálculo
                            st.session_state.calculo_solicitado = False
                            if 'df_final' in st.session_state:
                                del st.session_state.df_final
                            if 'df_com_aging' in st.session_state:
                                del st.session_state.df_com_aging

                            # Otimizar curva DI-PRE (vetorizado, sem loop por mes)
                            st.session_state.df_di_pre = otimizar_curva_di_pre(st.session_state.df_di_pre)

                            st.rerun()  # Recarregar a página para atualizar o estado
                        else:
                            st.error("❌ Não foi possível processar o arquivo CDI. Verifique o formato do arquivo.")
                            
                except Exception as e:
                    st.error(f"❌ Erro ao processar arquivo CDI: {str(e)}")
                    st.error("Verifique se o arquivo está no formato correto da BMF.")

    else:
        st.success("✅ **Dados CDI/DI-PRE carregados**")
        registros_cdi = len(st.session_state.df_di_pre)
        st.info(f"📊 {registros_cdi} registro(s) de CDI/DI-PRE disponível(eis)")
        
        # Mostrar botão para recarregar se necessário
        if st.button("🔄 Recarregar Dados CDI", key="recarregar_dados_cdi"):
            st.session_state.cdi_carregado = False
            st.session_state.calculo_solicitado = False
            if 'df_di_pre' in st.session_state:
                del st.session_state.df_di_pre
                st.rerun()  # Recarregar a página para atualizar o estado

            # Mostrar o head dos dados num expander
            if 'df_di_pre' in st.session_state:
                with st.expander("📊 Preview dos Dados CDI/DI-PRE", expanded=False):
                    st.dataframe(st.session_state.df_di_pre.head(10), use_container_width=True)

    st.markdown("---")
    
    # ETAPA 3: INFORMAÇÕES E CÁLCULO
    st.subheader("📊 3️⃣ Executar Cálculo")

    # ── Data Base ──────────────────────────────────────────────────────
    with st.container():
        st.markdown(
            """
            <style>
            div[data-testid="stDateInput"] { max-width: 220px; }
            </style>
            """,
            unsafe_allow_html=True,
        )
        col_date, col_info = st.columns([1, 2])
        with col_date:
            nova_data_base = st.date_input(
                "📅 Data Base",
                value=st.session_state.params.data_base_padrao,
                format="DD/MM/YYYY",
                key="input_data_base",
                help="Data de referência para correção monetária, aging e projeções.",
            )
        with col_info:
            st.markdown("")  # espaçamento vertical
            st.info(
                f"📅 **Data Base selecionada:** "
                f"**{nova_data_base.strftime('%d/%m/%Y')}**"
            )
        # Atualizar o parâmetro global
        st.session_state.params.data_base_padrao = datetime.combine(
            nova_data_base, datetime.min.time()
        )
        data_base_execucao = pd.Timestamp(st.session_state.params.data_base_padrao)

    st.markdown("---")

    # Seção de informações antes do cálculo
    st.write("**Informações do Processamento:**")

    col1, col2 = st.columns(2)

    with col1:
        st.metric("📊 Registros a Processar", f"{len(df_padronizado):,}")
    
    with col2:
        empresas_dados = df_padronizado['empresa'].nunique() if 'empresa' in df_padronizado.columns else 0
        st.metric("🏢 Empresas nos Dados", empresas_dados)

    
    # Verificar se todos os arquivos necessários estão carregados
    # TODAS as distribuidoras (incluindo VOLTZ) precisam de: índices, taxa de recuperação e CDI
    tem_todos_arquivos = tem_indices and tem_taxa_recuperacao and tem_cdi
    
    # Botão para calcular correção (SÓ APARECE SE TIVER TODOS OS ARQUIVOS)
    st.markdown("---")
    if tem_todos_arquivos:
        st.write("**✅ Todos os arquivos carregados! Agora você pode executar o cálculo:**")
        calculo_executado = st.button("💰 Calcular Correção Monetária Completa", type="primary", use_container_width=True, key="calcular_correcao_completa")
    else:
        arquivos_faltantes = []
        if not tem_indices:
            arquivos_faltantes.append("📊 Índices IGP-M/IPCA")
        if not tem_taxa_recuperacao:
            arquivos_faltantes.append("📈 Taxa de Recuperação")
        if not tem_cdi:
            arquivos_faltantes.append("📊 Dados CDI/DI-PRE")
        
        st.warning(f"⚠️ **Arquivos pendentes:** {', '.join(arquivos_faltantes)}")
        st.info("💡 Complete o carregamento de todos os arquivos acima para executar o cálculo.")
        calculo_executado = False
    
    if calculo_executado:
        # Marcar que o cálculo foi solicitado pelo usuário
        st.session_state.calculo_solicitado = True
        st.session_state.calculo_execucao_id = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        
        try:
            # ========== DASHBOARD DE PROGRESSO EM TEMPO REAL ==========
            st.markdown("### 🚀 **Processamento em Andamento**")
            
            
            # Barra de progresso principal
            progress_main = st.progress(0)
            status_text = st.empty()
            
            # Container para logs detalhados
            log_container = st.expander("📊 **Logs Detalhados de Performance**", expanded=False)

            # Pré-visualizações intermediárias por etapa (otimizadas para não sobrecarregar a UI)
            exibir_previews_etapas = st.checkbox(
                "📋 Exibir pré-visualizações por etapa",
                value=True,
                help="Mostra amostras parciais (tabelas enxutas) após etapas-chave do processamento.",
            )
            limite_preview = st.selectbox(
                "Registros por pré-visualização",
                options=[10, 20, 30, 50],
                index=1,
                disabled=not exibir_previews_etapas,
            )

            def exibir_preview_etapa(df_etapa, titulo_etapa, colunas_chave=None):
                """Exibe uma prévia parcial e enxuta do resultado intermediário por etapa."""
                if not exibir_previews_etapas:
                    return
                if df_etapa is None or df_etapa.empty:
                    return

                if colunas_chave is None:
                    colunas_chave = []

                colunas_base = [
                    'nome_cliente', 'contrato', 'empresa', 'data_base', 'data_vencimento',
                    'dias_atraso', 'aging', 'valor_liquido', 'valor_corrigido',
                ]

                colunas_ordenadas = []
                for col in colunas_base + colunas_chave:
                    if col in df_etapa.columns and col not in colunas_ordenadas:
                        colunas_ordenadas.append(col)

                if not colunas_ordenadas:
                    colunas_ordenadas = list(df_etapa.columns[:8])

                df_preview = df_etapa[colunas_ordenadas].head(limite_preview).copy()

                with st.expander(f"📋 Pré-visualização parcial - {titulo_etapa}", expanded=False):
                    st.caption(
                        f"Mostrando {len(df_preview)} de {len(df_etapa):,} registros "
                        f"com {len(colunas_ordenadas)} colunas-chave."
                    )
                    st.dataframe(df_preview, use_container_width=True, height=320)
            
            # Garantir que a data_base selecionada sobrescreva a coluna de entrada antes de qualquer cálculo
            df_entrada_calculo = df_padronizado.copy()
            df_entrada_calculo['data_base'] = data_base_execucao

            # Iniciar cronômetro
            import time
            inicio_processamento = time.time()
            total_registros = len(df_entrada_calculo)
            
            # ========== ETAPA 1: CÁLCULO DE AGING (20%) ==========
            status_text.text("🔄 Etapa 1/5: Calculando aging dos contratos...")
            progress_main.progress(0.1)
            
            with log_container:
                st.info(f"📊 **Iniciando processamento ultra-otimizado:** {total_registros:,} registros")
                
            etapa_inicio = time.time()
            df_com_aging = calc_aging.processar_aging_completo(df_entrada_calculo.copy())
            etapa_tempo = time.time() - etapa_inicio
            
            if df_com_aging.empty:
                st.error("❌ Erro ao calcular aging. Verifique os dados de entrada.")
                return
            
            progress_main.progress(0.2)
            velocidade_aging = total_registros / etapa_tempo if etapa_tempo > 0 else 0
            
            with log_container:
                st.success(f"✅ **Aging calculado:** {len(df_com_aging):,} registros em {etapa_tempo:.2f}s ({velocidade_aging:,.0f} reg/s)")

            exibir_preview_etapa(
                df_com_aging,
                "Etapa 1 - Aging",
                ['aging_taxa', 'valor_principal_limpo', 'valor_liquido'],
            )
            
            # ========== ETAPA 2: DETECÇÃO AUTOMÁTICA (40%) ==========
            status_text.text("� Etapa 2/5: Detectando tipo de distribuidora e regras...")
            progress_main.progress(0.3)
            
            # Obter nome do arquivo original (se disponível)
            nome_arquivo_original = "Distribuidora"  # Default
            if 'df_carregado' in st.session_state and st.session_state.df_carregado:
                primeiro_arquivo = list(st.session_state.df_carregado.keys())[0]
                nome_arquivo_original = primeiro_arquivo
            
            etapa_inicio = time.time()

            
            
            # Usar o novo método que detecta automaticamente VOLTZ vs Padrão
            df_final_temp = calc_correcao.processar_com_regras_especificas(
                df_com_aging.copy(), 
                nome_arquivo_original,  # Passa o nome do arquivo para detecção
                st.session_state.df_taxa_recuperacao
            )

            exibir_preview_etapa(
                df_final_temp,
                "Etapa 2 - Regras Específicas",
                ['taxa_recuperacao', 'prazo_recebimento', 'valor_recuperavel_ate_data_base'],
            )

            # Detectar se é VOLTZ para decidir o fluxo
            eh_voltz_detectado = 'VOLTZ' in nome_arquivo_original.upper()
            
            # ========== DEFINIR FUNÇÕES DE TRATAMENTO VOLTZ ==========
            def tratar_tipagem_dataframe_voltz(df):
                """
                Método que identifica e trata automaticamente todas as colunas e suas tipagens
                especificamente otimizado para dados VOLTZ
                """
                if df.empty:
                    return df
                
                df_tratado = df.copy()
                log_tipagem = []
                
                with st.expander("🔍 **Tratamento Automático de Tipagens - VOLTZ**", expanded=False):
                    st.info("🤖 Analisando e otimizando tipos de dados automaticamente...")
                    
                    # Contadores para estatísticas
                    colunas_numericas = 0
                    colunas_datas = 0
                    colunas_texto = 0
                    colunas_otimizadas = 0
                    
                    progress_tipagem = st.progress(0)
                    
                    for i, col in enumerate(df_tratado.columns):
                        try:
                            # Atualizar progress bar
                            progress_tipagem.progress((i + 1) / len(df_tratado.columns))
                            
                            # Amostrar dados para análise (para performance)
                            amostra = df_tratado[col].dropna().head(1000)
                            tipo_original = str(df_tratado[col].dtype)
                            
                            if amostra.empty:
                                log_tipagem.append(f"⚠️ {col}: Coluna vazia - mantida como {tipo_original}")
                                continue
                            
                            # 1. DETECTAR E TRATAR COLUNAS DE DATA
                            if any(keyword in col.lower() for keyword in ['data', 'date', 'vencimento', 'emissao', 'base']):
                                try:
                                    # Tentar múltiplos formatos de data
                                    df_tratado[col] = pd.to_datetime(df_tratado[col], 
                                                                   errors='coerce', 
                                                                   dayfirst=True,
                                                                   format='mixed')
                                    if not df_tratado[col].isna().all():
                                        colunas_datas += 1
                                        colunas_otimizadas += 1
                                        log_tipagem.append(f"📅 {col}: {tipo_original} → datetime64")
                                        continue
                                except:
                                    pass
                            
                            # 2. DETECTAR E TRATAR COLUNAS NUMÉRICAS
                            # Verificar se é numerico (incluindo valores com vírgula brasileira)
                            if df_tratado[col].dtype == 'object':
                                # Tentar conversão de números com vírgula brasileira
                                amostra_str = amostra.astype(str).str.replace(',', '.', regex=False)
                                
                                # Verificar se parece com número
                                pattern_numerico = r'^-?\d+(\.\d+)?$'
                                eh_numerico = amostra_str.str.match(pattern_numerico, na=False).sum() > (len(amostra) * 0.8)
                                
                                if eh_numerico:
                                    try:
                                        # Substituir vírgulas por pontos e converter
                                        df_tratado[col] = df_tratado[col].astype(str).str.replace(',', '.', regex=False)
                                        df_tratado[col] = pd.to_numeric(df_tratado[col], errors='coerce')
                                        
                                        # Otimizar tipo numérico
                                        if df_tratado[col].isna().sum() < len(df_tratado) * 0.1:  # Menos de 10% NaN
                                            # Verificar se são todos inteiros
                                            if df_tratado[col].dropna().apply(lambda x: x.is_integer()).all():
                                                # Otimizar tipo inteiro
                                                min_val = df_tratado[col].min()
                                                max_val = df_tratado[col].max()
                                                
                                                if min_val >= 0:
                                                    if max_val <= 255:
                                                        df_tratado[col] = df_tratado[col].astype('uint8')
                                                    elif max_val <= 65535:
                                                        df_tratado[col] = df_tratado[col].astype('uint16')
                                                    elif max_val <= 4294967295:
                                                        df_tratado[col] = df_tratado[col].astype('uint32')
                                                    else:
                                                        df_tratado[col] = df_tratado[col].astype('uint64')
                                                else:
                                                    if min_val >= -128 and max_val <= 127:
                                                        df_tratado[col] = df_tratado[col].astype('int8')
                                                    elif min_val >= -32768 and max_val <= 32767:
                                                        df_tratado[col] = df_tratado[col].astype('int16')
                                                    elif min_val >= -2147483648 and max_val <= 2147483647:
                                                        df_tratado[col] = df_tratado[col].astype('int32')
                                                    else:
                                                        df_tratado[col] = df_tratado[col].astype('int64')
                                            else:
                                                # Usar float32 se possível, senão float64
                                                df_tratado[col] = df_tratado[col].astype('float32')
                                        
                                        colunas_numericas += 1
                                        colunas_otimizadas += 1
                                        tipo_novo = str(df_tratado[col].dtype)
                                        log_tipagem.append(f"🔢 {col}: {tipo_original} → {tipo_novo}")
                                        continue
                                    except:
                                        pass
                            
                            # 3. OTIMIZAR COLUNAS DE TEXTO
                            if df_tratado[col].dtype == 'object':
                                # Verificar se pode ser categoria
                                unique_ratio = df_tratado[col].nunique() / len(df_tratado)
                                
                                if unique_ratio < 0.1 and df_tratado[col].nunique() < 1000:  # Baixa cardinalidade
                                    df_tratado[col] = df_tratado[col].astype('category')
                                    colunas_otimizadas += 1
                                    log_tipagem.append(f"📂 {col}: object → category (cardinalidade: {df_tratado[col].nunique()})")
                                else:
                                    # Tentar string mais eficiente
                                    try:
                                        df_tratado[col] = df_tratado[col].astype('string')
                                        log_tipagem.append(f"📝 {col}: object → string")
                                    except:
                                        log_tipagem.append(f"📝 {col}: mantido como object")
                                
                                colunas_texto += 1
                            
                            # 4. OTIMIZAR COLUNAS BOOLEAN
                            if df_tratado[col].dtype == 'bool' or amostra.isin([True, False, 0, 1, 'True', 'False', 'true', 'false']).all():
                                try:
                                    df_tratado[col] = df_tratado[col].astype('bool')
                                    colunas_otimizadas += 1
                                    log_tipagem.append(f"✅ {col}: → bool")
                                except:
                                    pass
                            
                        except Exception as e:
                            log_tipagem.append(f"❌ {col}: Erro na tipagem - {str(e)[:50]}...")
                            continue
                    
                    # Estatísticas finais
                    col1, col2, col3, col4 = st.columns(4)
                    
                    with col1:
                        st.metric("🔢 Numéricas", colunas_numericas, "otimizadas")
                    with col2:
                        st.metric("📅 Datas", colunas_datas, "convertidas")
                    with col3:
                        st.metric("📝 Texto", colunas_texto, "categorizadas")
                    with col4:
                        st.metric("⚡ Otimizadas", colunas_otimizadas, f"de {len(df_tratado.columns)}")
                    
                    # Comparar tamanhos
                    memoria_original = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
                    memoria_otimizada = df_tratado.memory_usage(deep=True).sum() / 1024 / 1024  # MB
                    reducao_percentual = ((memoria_original - memoria_otimizada) / memoria_original) * 100
                    
                    st.success(f"✅ **Otimização completa!** Redução de memória: {reducao_percentual:.1f}% ({memoria_original:.1f}MB → {memoria_otimizada:.1f}MB)")
                
                return df_tratado
            
            def salvar_dados_tratados_voltz(df_tipado):
                """
                Método que cria botão para salvar os dados tratados na pasta 'data'
                """
                # Garantir que a pasta 'data' existe
                data_path = os.path.join(os.getcwd(), 'data')
                if not os.path.exists(data_path):
                    os.makedirs(data_path)
                    st.info(f"📁 Pasta 'data' criada em: {data_path}")
                
                # Interface para salvar
                st.subheader("💾 **Exportar Dados VOLTZ Tratados**")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Opções de formato
                    formato_export = st.selectbox(
                        "Formato de exportação:",
                        ["Excel (.xlsx)", "CSV (.csv)", "Pickle (.pkl)"],
                        help="Escolha o formato para salvar os dados tratados"
                    )
                    
                    # Nome do arquivo
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    nome_arquivo_default = f"VOLTZ_Dados_Tratados_{timestamp}"
                    
                    nome_arquivo = st.text_input(
                        "Nome do arquivo:",
                        value=nome_arquivo_default,
                        help="Nome do arquivo (sem extensão)"
                    )
                
                with col2:
                    # Informações sobre o arquivo
                    st.info(f"📊 **Registros:** {len(df_tipado):,}")
                    st.info(f"📋 **Colunas:** {len(df_tipado.columns)}")
                    memoria_mb = df_tipado.memory_usage(deep=True).sum() / 1024 / 1024
                    st.info(f"💾 **Memória:** {memoria_mb:.1f} MB")
                
                # Botão de exportação
                if st.button("🚀 **Salvar Dados Tratados**", type="primary"):
                    try:
                        # Determinar extensão e caminho
                        if formato_export == "Excel (.xlsx)":
                            extensao = ".xlsx"
                            caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                            
                            with st.spinner("📤 Salvando arquivo Excel..."):
                                df_tipado.to_excel(caminho_arquivo, index=False)
                                
                        elif formato_export == "CSV (.csv)":
                            extensao = ".csv"
                            caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                            
                            with st.spinner("📤 Salvando arquivo CSV..."):
                                salvar_csv_brasil(df_tipado, caminho_arquivo, casas_decimais=4)
                                
                        elif formato_export == "Pickle (.pkl)":
                            extensao = ".pkl"
                            caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                            
                            with st.spinner("📤 Salvando arquivo Pickle..."):
                                df_tipado.to_pickle(caminho_arquivo)
                        
                        # Verificar se arquivo foi criado
                        if os.path.exists(caminho_arquivo):
                            tamanho_arquivo = os.path.getsize(caminho_arquivo) / 1024 / 1024  # MB
                            
                            st.success(f"✅ **Arquivo salvo com sucesso!**")
                            st.info(f"📁 **Localização:** `{caminho_arquivo}`")
                            st.info(f"📊 **Tamanho:** {tamanho_arquivo:.1f} MB")
                            st.info(f"⏰ **Timestamp:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
                            
                            # Balões de comemoração
                            st.balloons()
                            
                            # Log no session_state para histórico
                            if 'historico_exports_voltz' not in st.session_state:
                                st.session_state.historico_exports_voltz = []
                            
                            st.session_state.historico_exports_voltz.append({
                                'timestamp': datetime.now(),
                                'arquivo': caminho_arquivo,
                                'formato': formato_export,
                                'registros': len(df_tipado),
                                'tamanho_mb': tamanho_arquivo
                            })
                            
                        else:
                            st.error("❌ Erro: Arquivo não foi criado corretamente")
                            
                    except Exception as e:
                        st.error(f"❌ **Erro ao salvar arquivo:** {str(e)}")
                        st.exception(e)
                
            # Para VOLTZ: o cálculo já está finalizado
            if eh_voltz_detectado:
                with log_container:
                    st.success("⚡ **VOLTZ detectada:** Cálculo finalizado no processamento específico!")
                
                # Para VOLTZ, apenas aplicar o valor justo reajustado
                df_final_temp = calc_correcao.calcular_valor_justo_reajustado(df_final_temp)

                exibir_preview_etapa(
                    df_final_temp,
                    "Etapa Final VOLTZ",
                    [
                        'valor_corrigido_ate_recebimento', 'valor_recuperavel_ate_recebimento',
                        'remuneracao_variavel_perc', 'remuneracao_variavel_valor',
                        'valor_recebimento_pos_rv', 'valor_justo'
                    ],
                )
                
                # Finalizar processamento para VOLTZ
                st.session_state.df_final = df_final_temp
                st.session_state.df_com_aging = df_com_aging

                caminho_exportado, novo_arquivo = exportar_resultado_final_excel(
                    st.session_state.df_final,
                    eh_voltz=True,
                )
                if caminho_exportado:
                    if novo_arquivo:
                        st.success(f"💾 Resultado final VOLTZ exportado automaticamente em: {caminho_exportado}")
                    else:
                        st.info(f"💾 Última exportação automática VOLTZ: {caminho_exportado}")
                
                progress_main.progress(1.0)
                etapa_tempo = time.time() - etapa_inicio
                tempo_total = time.time() - inicio_processamento
                
                status_text.text("✅ Processamento VOLTZ concluído com sucesso!")
                
                # Métricas finais para VOLTZ
                velocidade_total = len(df_final_temp) / tempo_total if tempo_total > 0 else 0
                
                metric_col1, metric_col2, metric_col3 = st.columns(3)
                
                with metric_col1:
                    st.metric("📊 Registros", f"{len(df_final_temp):,}", "✅ VOLTZ")
                with metric_col2:
                    st.metric("⚡ Velocidade", f"{velocidade_total:,.0f}", "registros/seg")
                with metric_col3:
                    st.metric("⏱️ Tempo Total", f"{tempo_total:.1f}s", "🎯 Ultra-rápido")
                
                with log_container:
                    st.balloons()
                    st.success("🎉 **PROCESSAMENTO VOLTZ ULTRA-OTIMIZADO CONCLUÍDO!**")
                    st.info(f"⚡ **VOLTZ Performance:** {len(df_final_temp):,} registros em {tempo_total:.2f}s")
                    st.info(f"🚀 **Throughput VOLTZ:** {velocidade_total:,.0f} registros/segundo")
                

                # ========== TRATAMENTO AUTOMÁTICO DE TIPAGENS VOLTZ ==========
                def tratar_tipagem_dataframe_voltz(df):
                    """
                    Método que identifica e trata automaticamente todas as colunas e suas tipagens
                    especificamente otimizado para dados VOLTZ
                    """
                    if df.empty:
                        return df
                    
                    df_tratado = df.copy()
                    log_tipagem = []
                    
                    with st.expander("🔍 **Tratamento Automático de Tipagens - VOLTZ**", expanded=False):
                        st.info("🤖 Analisando e otimizando tipos de dados automaticamente...")
                        
                        # Contadores para estatísticas
                        colunas_numericas = 0
                        colunas_datas = 0
                        colunas_texto = 0
                        colunas_otimizadas = 0
                        
                        progress_tipagem = st.progress(0)
                        
                        for i, col in enumerate(df_tratado.columns):
                            try:
                                # Atualizar progress bar
                                progress_tipagem.progress((i + 1) / len(df_tratado.columns))
                                
                                # Amostrar dados para análise (para performance)
                                amostra = df_tratado[col].dropna().head(1000)
                                tipo_original = str(df_tratado[col].dtype)
                                
                                if amostra.empty:
                                    log_tipagem.append(f"⚠️ {col}: Coluna vazia - mantida como {tipo_original}")
                                    continue
                                
                                # 1. DETECTAR E TRATAR COLUNAS DE DATA
                                if any(keyword in col.lower() for keyword in ['data', 'date', 'vencimento', 'emissao', 'base']):
                                    try:
                                        # Tentar múltiplos formatos de data
                                        df_tratado[col] = pd.to_datetime(df_tratado[col], 
                                                                       errors='coerce', 
                                                                       dayfirst=True,
                                                                       format='mixed')
                                        if not df_tratado[col].isna().all():
                                            colunas_datas += 1
                                            colunas_otimizadas += 1
                                            log_tipagem.append(f"📅 {col}: {tipo_original} → datetime64")
                                            continue
                                    except:
                                        pass
                                
                                # 2. DETECTAR E TRATAR COLUNAS NUMÉRICAS
                                # Verificar se é numerico (incluindo valores com vírgula brasileira)
                                if df_tratado[col].dtype == 'object':
                                    # Tentar conversão de números com vírgula brasileira
                                    amostra_str = amostra.astype(str).str.replace(',', '.', regex=False)
                                    
                                    # Verificar se parece com número
                                    pattern_numerico = r'^-?\d+(\.\d+)?$'
                                    eh_numerico = amostra_str.str.match(pattern_numerico, na=False).sum() > (len(amostra) * 0.8)
                                    
                                    if eh_numerico:
                                        try:
                                            # Substituir vírgulas por pontos e converter
                                            df_tratado[col] = df_tratado[col].astype(str).str.replace(',', '.', regex=False)
                                            df_tratado[col] = pd.to_numeric(df_tratado[col], errors='coerce')
                                            
                                            # Otimizar tipo numérico
                                            if df_tratado[col].isna().sum() < len(df_tratado) * 0.1:  # Menos de 10% NaN
                                                # Verificar se são todos inteiros
                                                if df_tratado[col].dropna().apply(lambda x: x.is_integer()).all():
                                                    # Otimizar tipo inteiro
                                                    min_val = df_tratado[col].min()
                                                    max_val = df_tratado[col].max()
                                                    
                                                    if min_val >= 0:
                                                        if max_val <= 255:
                                                            df_tratado[col] = df_tratado[col].astype('uint8')
                                                        elif max_val <= 65535:
                                                            df_tratado[col] = df_tratado[col].astype('uint16')
                                                        elif max_val <= 4294967295:
                                                            df_tratado[col] = df_tratado[col].astype('uint32')
                                                        else:
                                                            df_tratado[col] = df_tratado[col].astype('uint64')
                                                    else:
                                                        if min_val >= -128 and max_val <= 127:
                                                            df_tratado[col] = df_tratado[col].astype('int8')
                                                        elif min_val >= -32768 and max_val <= 32767:
                                                            df_tratado[col] = df_tratado[col].astype('int16')
                                                        elif min_val >= -2147483648 and max_val <= 2147483647:
                                                            df_tratado[col] = df_tratado[col].astype('int32')
                                                        else:
                                                            df_tratado[col] = df_tratado[col].astype('int64')
                                                else:
                                                    # Usar float32 se possível, senão float64
                                                    df_tratado[col] = df_tratado[col].astype('float32')
                                            
                                            colunas_numericas += 1
                                            colunas_otimizadas += 1
                                            tipo_novo = str(df_tratado[col].dtype)
                                            log_tipagem.append(f"🔢 {col}: {tipo_original} → {tipo_novo}")
                                            continue
                                        except:
                                            pass
                                
                                # 3. OTIMIZAR COLUNAS DE TEXTO
                                if df_tratado[col].dtype == 'object':
                                    # Verificar se pode ser categoria
                                    unique_ratio = df_tratado[col].nunique() / len(df_tratado)
                                    
                                    if unique_ratio < 0.1 and df_tratado[col].nunique() < 1000:  # Baixa cardinalidade
                                        df_tratado[col] = df_tratado[col].astype('category')
                                        colunas_otimizadas += 1
                                        log_tipagem.append(f"📂 {col}: object → category (cardinalidade: {df_tratado[col].nunique()})")
                                    else:
                                        # Tentar string mais eficiente
                                        try:
                                            df_tratado[col] = df_tratado[col].astype('string')
                                            log_tipagem.append(f"📝 {col}: object → string")
                                        except:
                                            log_tipagem.append(f"📝 {col}: mantido como object")
                                    
                                    colunas_texto += 1
                                
                                # 4. OTIMIZAR COLUNAS BOOLEAN
                                if df_tratado[col].dtype == 'bool' or amostra.isin([True, False, 0, 1, 'True', 'False', 'true', 'false']).all():
                                    try:
                                        df_tratado[col] = df_tratado[col].astype('bool')
                                        colunas_otimizadas += 1
                                        log_tipagem.append(f"✅ {col}: → bool")
                                    except:
                                        pass
                                
                            except Exception as e:
                                log_tipagem.append(f"❌ {col}: Erro na tipagem - {str(e)[:50]}...")
                                continue
                        
                        # Estatísticas finais
                        col1, col2, col3, col4 = st.columns(4)
                        
                        with col1:
                            st.metric("🔢 Numéricas", colunas_numericas, "otimizadas")
                        with col2:
                            st.metric("📅 Datas", colunas_datas, "convertidas")
                        with col3:
                            st.metric("📝 Texto", colunas_texto, "categorizadas")
                        with col4:
                            st.metric("⚡ Otimizadas", colunas_otimizadas, f"de {len(df_tratado.columns)}")
                        
                        # Comparar tamanhos
                        memoria_original = df.memory_usage(deep=True).sum() / 1024 / 1024  # MB
                        memoria_otimizada = df_tratado.memory_usage(deep=True).sum() / 1024 / 1024  # MB
                        reducao_percentual = ((memoria_original - memoria_otimizada) / memoria_original) * 100
                        
                        st.success(f"✅ **Otimização completa!** Redução de memória: {reducao_percentual:.1f}% ({memoria_original:.1f}MB → {memoria_otimizada:.1f}MB)")
                    
                    return df_tratado
                
                # Aplicar tratamento de tipagem no dataframe VOLTZ
                st.info("🔧 **Aplicando tratamento automático de tipagens...**")
                df_final_temp_tipado = tratar_tipagem_dataframe_voltz(df_final_temp)
                
                # Atualizar no session_state
                st.session_state.df_final_voltz_tipado = df_final_temp_tipado
                
                # ========== BOTÃO PARA SALVAR DADOS TRATADOS ==========
                def salvar_dados_tratados_voltz(df_tipado):
                    """
                    Método que cria botão para salvar os dados tratados na pasta 'data'
                    """
                    # Garantir que a pasta 'data' existe
                    data_path = os.path.join(os.getcwd(), 'data')
                    if not os.path.exists(data_path):
                        os.makedirs(data_path)
                        st.info(f"📁 Pasta 'data' criada em: {data_path}")
                    
                    # Interface para salvar
                    st.subheader("💾 **Exportar Dados VOLTZ Tratados**")
                    
                    col1, col2 = st.columns(2)
                    
                    with col1:
                        # Opções de formato
                        formato_export = st.selectbox(
                            "Formato de exportação:",
                            ["Excel (.xlsx)", "CSV (.csv)", "Pickle (.pkl)"],
                            help="Escolha o formato para salvar os dados tratados"
                        )
                        
                        # Nome do arquivo
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        nome_arquivo_default = f"VOLTZ_Dados_Tratados_{timestamp}"
                        
                        nome_arquivo = st.text_input(
                            "Nome do arquivo:",
                            value=nome_arquivo_default,
                            help="Nome do arquivo (sem extensão)"
                        )
                    
                    with col2:
                        # Informações sobre o arquivo
                        st.info(f"📊 **Registros:** {len(df_tipado):,}")
                        st.info(f"📋 **Colunas:** {len(df_tipado.columns)}")
                        memoria_mb = df_tipado.memory_usage(deep=True).sum() / 1024 / 1024
                        st.info(f"💾 **Memória:** {memoria_mb:.1f} MB")
                    
                    # Botão de exportação
                    if st.button("🚀 **Salvar Dados Tratados**", type="primary"):
                        try:
                            # Determinar extensão e caminho
                            if formato_export == "Excel (.xlsx)":
                                extensao = ".xlsx"
                                caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                                
                                with st.spinner("📤 Salvando arquivo Excel..."):
                                    df_tipado.to_excel(caminho_arquivo, index=False)
                                    
                            elif formato_export == "CSV (.csv)":
                                extensao = ".csv"
                                caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                                
                                with st.spinner("📤 Salvando arquivo CSV..."):
                                    salvar_csv_brasil(df_tipado, caminho_arquivo, casas_decimais=4)
                                    
                            elif formato_export == "Pickle (.pkl)":
                                extensao = ".pkl"
                                caminho_arquivo = os.path.join(data_path, f"{nome_arquivo}{extensao}")
                                
                                with st.spinner("📤 Salvando arquivo Pickle..."):
                                    df_tipado.to_pickle(caminho_arquivo)
                            
                            # Verificar se arquivo foi criado
                            if os.path.exists(caminho_arquivo):
                                tamanho_arquivo = os.path.getsize(caminho_arquivo) / 1024 / 1024  # MB
                                
                                st.success(f"✅ **Arquivo salvo com sucesso!**")
                                st.info(f"📁 **Localização:** `{caminho_arquivo}`")
                                st.info(f"📊 **Tamanho:** {tamanho_arquivo:.1f} MB")
                                st.info(f"⏰ **Timestamp:** {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
                                
                                # Balões de comemoração
                                st.balloons()
                                
                                # Log no session_state para histórico
                                if 'historico_exports_voltz' not in st.session_state:
                                    st.session_state.historico_exports_voltz = []
                                
                                st.session_state.historico_exports_voltz.append({
                                    'timestamp': datetime.now(),
                                    'arquivo': caminho_arquivo,
                                    'formato': formato_export,
                                    'registros': len(df_tipado),
                                    'tamanho_mb': tamanho_arquivo
                                })
                                
                            else:
                                st.error("❌ Erro: Arquivo não foi criado corretamente")
                                
                        except Exception as e:
                            st.error(f"❌ **Erro ao salvar arquivo:** {str(e)}")
                            st.exception(e)

                
                # Chamar função de salvamento
                salvar_dados_tratados_voltz(df_final_temp_tipado)

                # Para VOLTZ, terminar aqui
                return
            else:
                etapa_tempo = time.time() - etapa_inicio
                progress_main.progress(0.4)
                
                # Detectar se é VOLTZ para mostrar no log
                eh_voltz_detectado = 'VOLTZ' in nome_arquivo_original.upper()
                with log_container:
                    if eh_voltz_detectado:
                        st.info("🎯 **Sistema VOLTZ detectado**")
                    else:
                        st.info("🎯 **Sistema padrão detectado**")
                    
                # ========== ETAPA 3: ÍNDICES CUSTOMIZADOS (60%) ==========
                status_text.text("📈 Etapa 3/5: Aplicando índices econômicos customizados...")
                progress_main.progress(0.5)
                
                # ========== VERIFICAR SE TEMOS ÍNDICES CUSTOMIZADOS ==========
                tem_indices_economicos = 'df_indices_economicos' in st.session_state and not st.session_state.df_indices_economicos.empty
                tem_indices_igpm = 'df_indices_igpm' in st.session_state and not st.session_state.df_indices_igpm.empty
                
                if tem_indices_economicos or tem_indices_igpm:
                    etapa_inicio = time.time()
                    
                    with log_container:
                        st.info("📊 **Iniciando merge vetorizado** de índices temporais...")

                    # ========== SELEÇÃO INTELIGENTE DOS ÍNDICES ==========
                    # Detectar se é VOLTZ para usar índices corretos
                    if eh_voltz_detectado and tem_indices_igpm:
                        # VOLTZ: Usar aba IGPM
                        df_indices = st.session_state.df_indices_igpm.copy()
                        tipo_indice = "IGPM (VOLTZ)"
                        with log_container:
                            st.info("⚡ **VOLTZ detectada:** Usando índices da aba IGPM")
                    elif tem_indices_economicos:
                        # DISTRIBUIDORAS: Usar aba IGPM_IPCA
                        df_indices = st.session_state.df_indices_economicos.copy()
                        tipo_indice = "IGPM_IPCA (Distribuidoras)"
                        with log_container:
                            st.info("🏢 **Distribuidora padrão:** Usando índices da aba IGPM_IPCA")
                    else:
                        # Fallback: usar qualquer índice disponível
                        if tem_indices_igpm:
                            df_indices = st.session_state.df_indices_igpm.copy()
                            tipo_indice = "IGPM (Fallback)"
                        else:
                            df_indices = st.session_state.df_indices_economicos.copy()
                            tipo_indice = "IGPM_IPCA (Fallback)"
                        
                        with log_container:
                            st.warning(f"⚠️ **Fallback:** Usando {tipo_indice}")
                    
                    # ========== PROCESSAR ÍNDICES SELECIONADOS ==========
                    with log_container:
                        registros_indices = len(df_indices)
                        periodo_min = df_indices['data'].min().strftime('%Y-%m')
                        periodo_max = df_indices['data'].max().strftime('%Y-%m')
                        st.success(f"✅ **Índices carregados:** {tipo_indice} - {registros_indices:,} registros ({periodo_min} a {periodo_max})")
                    
                    df_indices['data'] = pd.to_datetime(df_indices['data'])
                    df_indices = df_indices.sort_values('data')
                    
                    # Criar índices com mês anterior para cálculo da taxa mensal
                    df_indices['data_mes_anterior'] = df_indices['data'].shift(1)
                    df_indices['indice_mes_anterior'] = df_indices['indice'].shift(1)
                    
                    # Calcular taxa mensal = indice_atual - indice_anterior (diferença simples)
                    df_indices['taxa_mensal'] = 1 - df_indices['indice_mes_anterior'] / df_indices['indice']
                    df_indices['taxa_diaria'] = (df_indices['taxa_mensal'] + 1) ** (1/30) - 1

                    # Preparar DataFrame principal
                    df_final_temp = df_final_temp.copy()
                    df_final_temp['data_vencimento_limpa'] = pd.to_datetime(df_final_temp['data_vencimento_limpa'], errors='coerce')
                    df_final_temp['data_base'] = pd.to_datetime(df_final_temp['data_base'], errors='coerce')

                    progress_main.progress(0.52)
                    
                    # ==== MERGE 1: DATA BASE (ULTRA-OTIMIZADO) ====
                    with log_container:
                        st.info("🔄 **Merge 1/2:** Índices da data base (O(log n))...")
                    
                    # Criar coluna auxiliar para merge (ano-mês)
                    df_indices['ano_mes'] = df_indices['data'].dt.to_period('M')
                    df_final_temp['ano_mes_base'] = df_final_temp['data_base'].dt.to_period('M')
                    df_final_temp['ano_mes_venc'] = df_final_temp['data_vencimento_limpa'].dt.to_period('M')

                    registros_antes_merge_indices = len(df_final_temp)

                    # Garantir cardinalidade 1:1 por ano_mes para evitar explosão de linhas no merge.
                    df_indices_merge = df_indices[
                        ['ano_mes', 'indice', 'indice_mes_anterior', 'taxa_mensal', 'taxa_diaria']
                    ].dropna(subset=['ano_mes']).copy()

                    duplicatas_ano_mes = int(df_indices_merge.duplicated(subset=['ano_mes']).sum())
                    if duplicatas_ano_mes > 0:
                        with log_container:
                            st.warning(
                                f"⚠️ Foram encontradas {duplicatas_ano_mes:,} linhas duplicadas de índices por mês. "
                                "Aplicando deduplicação para manter 1 registro por ano-mês."
                            )
                        df_indices_merge = (
                            df_indices_merge
                            .sort_values('ano_mes')
                            .drop_duplicates(subset=['ano_mes'], keep='last')
                            .reset_index(drop=True)
                        )
                    
                    # Merge com data base
                    df_merged_base = df_final_temp.merge(
                        df_indices_merge.rename(columns={
                            'ano_mes': 'ano_mes_base',
                            'indice': 'indice_mes_base',
                            'indice_mes_anterior': 'indice_mes_anterior_base',
                            'taxa_mensal': 'taxa_mensal_base',
                            'taxa_diaria': 'taxa_diaria_base'
                        }),
                        on='ano_mes_base',
                        how='left',
                        validate='m:1'
                    )

                    if len(df_merged_base) != registros_antes_merge_indices:
                        with log_container:
                            st.warning(
                                f"⚠️ Merge de data base alterou contagem de linhas "
                                f"({registros_antes_merge_indices:,} → {len(df_merged_base):,})."
                            )
                    
                    progress_main.progress(0.54)
                    
                    # ==== MERGE 2: DATA VENCIMENTO (ULTRA-OTIMIZADO) ====
                    with log_container:
                        st.info("🔄 **Merge 2/2:** Índices da data vencimento (O(log n))...")
                        
                    df_merged_completo = df_merged_base.merge(
                        df_indices_merge.rename(columns={
                            'ano_mes': 'ano_mes_venc',
                            'indice': 'indice_mes_venc',
                            'indice_mes_anterior': 'indice_mes_anterior_venc',
                            'taxa_mensal': 'taxa_mensal_venc',
                            'taxa_diaria': 'taxa_diaria_venc'
                        }),
                        on='ano_mes_venc',
                        how='left',
                        validate='m:1'
                    )

                    if len(df_merged_completo) != registros_antes_merge_indices:
                        with log_container:
                            st.warning(
                                f"⚠️ Merge de data vencimento alterou contagem de linhas "
                                f"({registros_antes_merge_indices:,} → {len(df_merged_completo):,})."
                            )

                    # ==== FALLBACK: ÚLTIMO ÍNDICE DISPONÍVEL ==== 
                    # Quando a competência da data_base/data_vencimento não existe na curva,
                    # usar o último índice válido disponível para evitar fator=1 silencioso.
                    serie_indices_validos = pd.to_numeric(df_indices_merge.get('indice'), errors='coerce').dropna()
                    if not serie_indices_validos.empty:
                        ultimo_indice_disponivel = float(serie_indices_validos.iloc[-1])

                        mask_base_sem_indice = df_merged_completo['indice_mes_base'].isna()
                        qtd_base_sem_indice = int(mask_base_sem_indice.sum())
                        if qtd_base_sem_indice > 0:
                            df_merged_completo.loc[mask_base_sem_indice, 'indice_mes_base'] = ultimo_indice_disponivel
                            with log_container:
                                st.warning(
                                    f"⚠️ {qtd_base_sem_indice:,} registro(s) sem índice da data base. "
                                    f"Aplicado fallback para o último índice disponível ({ultimo_indice_disponivel:.6f})."
                                )

                        mask_venc_sem_indice = df_merged_completo['indice_mes_venc'].isna()
                        qtd_venc_sem_indice = int(mask_venc_sem_indice.sum())
                        if qtd_venc_sem_indice > 0:
                            df_merged_completo.loc[mask_venc_sem_indice, 'indice_mes_venc'] = ultimo_indice_disponivel
                            with log_container:
                                st.warning(
                                    f"⚠️ {qtd_venc_sem_indice:,} registro(s) sem índice do vencimento. "
                                    f"Aplicado fallback para o último índice disponível ({ultimo_indice_disponivel:.6f})."
                                )
                    
                    progress_main.progress(0.56)
                    
                    # ==== CÁLCULO DOS ÍNDICES DIÁRIOS (VETORIZADO) ====
                    with log_container:
                        st.info("🧮 **Cálculo vetorizado** de índices diários...")
                    
                    # Aplicar cálculo vetorizado
                    with log_container:
                        st.info("🔄 **Aplicando** índices data base...")
                    df_merged_completo['indice_base_diario'] = calcular_indice_diario_vetorizado(
                        df=df_merged_completo,
                        coluna_data='data_base',
                        coluna_indice_mes='indice_mes_base',
                        coluna_indice_mes_anterior='indice_mes_anterior_base',
                        coluna_taxa_mensal='taxa_mensal_base',
                        coluna_taxa_diaria='taxa_diaria_base',
                    )
                    
                    progress_main.progress(0.58)
                    
                    with log_container:
                        st.info("🔄 **Aplicando** índices data vencimento...")
                    df_merged_completo['indice_venc_diario'] = calcular_indice_diario_vetorizado(
                        df=df_merged_completo,
                        coluna_data='data_vencimento_limpa',
                        coluna_indice_mes='indice_mes_venc',
                        coluna_indice_mes_anterior='indice_mes_anterior_venc',
                        coluna_taxa_mensal='taxa_mensal_venc',
                        coluna_taxa_diaria='taxa_diaria_venc',
                    )
                    
                    # ==== CÁLCULO DO FATOR DE CORREÇÃO (ULTRA-RÁPIDO) ====
                    # Mask para registros válidos
                    mask_validos = (
                        df_merged_completo['indice_base_diario'].notna()
                        & df_merged_completo['indice_venc_diario'].notna()
                        & (df_merged_completo['indice_base_diario'] > 0)
                        & (df_merged_completo['indice_venc_diario'] > 0)
                    )
                    
                    # Fator de correção = indice_vencimento / indice_base
                    df_merged_completo['fator_correcao'] = 1.0  # Default
                    df_merged_completo.loc[mask_validos, 'fator_correcao'] = (
                        df_merged_completo.loc[mask_validos, 'indice_base_diario'] /
                        df_merged_completo.loc[mask_validos, 'indice_venc_diario']
                    )
                    
                    # ==== APLICAR CORREÇÃO MONETÁRIA (VETORIZADA) ====
                    df_merged_completo = aplicar_correcao_monetaria_vetorizada(
                        df_merged_completo,
                        coluna_fator='fator_correcao',
                    )

                    # Atualizar colunas finais
                    df_merged_completo.loc[mask_validos, 'indice_vencimento'] = df_merged_completo.loc[mask_validos, 'indice_venc_diario']
                    df_merged_completo.loc[mask_validos, 'indice_base'] = df_merged_completo.loc[mask_validos, 'indice_base_diario']

                    registros_customizados = mask_validos.sum()
                    total_registros = len(df_merged_completo)
                    percentual = (registros_customizados / total_registros) * 100

                    # Limpar colunas auxiliares
                    colunas_temp = [
                        'ano_mes_base', 'ano_mes_venc', 'indice_mes_base', 'indice_mes_venc',
                        'taxa_mensal_base', 'taxa_mensal_venc', 'data_fechamento_base', 'data_fechamento_venc',
                        'indice_base_diario', 'indice_venc_diario', 'indice_mes_anterior_base', 'taxa_diaria_base', 'indice_mes_anterior_venc', 'taxa_diaria_venc'
                    ]
                    df_final_temp = df_merged_completo.drop(columns=[col for col in colunas_temp if col in df_merged_completo.columns])
                    
                    etapa_tempo = time.time() - etapa_inicio
                    velocidade_indices = registros_customizados / etapa_tempo if etapa_tempo > 0 else 0
                    
                    with log_container:
                        st.success(f"✅ **Índices customizados aplicados:** {registros_customizados:,}/{total_registros:,} registros ({percentual:.1f}%) em {etapa_tempo:.2f}s ({velocidade_indices:,.0f} reg/s)")
                    
                else:
                    with log_container:
                        st.info("ℹ️ **Usando correção padrão** do sistema (IGPM/IPCA automático)")

                
                # ========== ETAPA 4: CÁLCULO DE CORREÇÃO MONETÁRIA FINAL (80%) ==========
                status_text.text("� Etapa 4/5: Calculando correção monetária final...")
                progress_main.progress(0.7)
                
                etapa_inicio = time.time()
                
                if df_final_temp.empty:
                    st.error("❌ Erro ao processar correção monetária.")
                    return

                # ==== APLICAR CORREÇÃO MONETÁRIA FINAL (VETORIZADA) ====
                with log_container:
                    st.info("💰 **Calculando correção monetária** vetorizada...")

                df_final_temp = aplicar_correcao_monetaria_vetorizada(
                    df_final_temp,
                    coluna_fator='fator_correcao',
                )

                # Renomear fator_correcao para fator_correcao_ate_data_base
                df_final_temp.rename(columns={'fator_correcao': 'fator_correcao_ate_data_base'}, inplace=True)

                progress_main.progress(0.75)
                
                etapa_tempo = time.time() - etapa_inicio
                velocidade_correcao_final = len(df_final_temp) / etapa_tempo if etapa_tempo > 0 else 0
                
                with log_container:
                    st.success(f"✅ **Correção monetária final:** {len(df_final_temp):,} registros em {etapa_tempo:.2f}s ({velocidade_correcao_final:,.0f} reg/s)")

                exibir_preview_etapa(
                    df_final_temp,
                    "Etapa 4 - Correção Monetária",
                    [
                        'fator_correcao_ate_data_base', 'correcao_monetaria',
                        'juros_moratorios', 'multa', 'valor_corrigido'
                    ],
                )

                # return
                # ========== ETAPA 5: CÁLCULO DO VALOR JUSTO PARA DISTRIBUIDORAS ==========
                # APENAS para distribuidoras padrão (não-VOLTZ)
                status_text.text("⚖️ Etapa 5/5: Calculando valor justo para distribuidoras...")
                progress_main.progress(0.8)
                
                etapa_inicio = time.time()
                
                with log_container:
                    st.info("⚖️ **Iniciando cálculo de valor justo** para distribuidoras padrão...")

                try:
                    # ========== USAR MÓDULO ESPECÍFICO PARA DISTRIBUIDORAS ==========
                    calc_valor_justo_dist = CalculadorValorJustoDistribuidoras(st.session_state.params)
                    
                    # Processar valor justo completo para distribuidoras
                    df_final_temp = calc_valor_justo_dist.processar_valor_justo_distribuidoras(
                        df_final_temp, 
                        log_container, 
                        progress_main
                    )
                    # ============= ETAPA 7: CALCULAR VALOR JUSTO REAJUSTADO =============
                    # Aplicar descontos por aging sobre o valor justo
                    df_final_temp = calc_correcao.calcular_valor_justo_reajustado(df_final_temp)

                    exibir_preview_etapa(
                        df_final_temp,
                        "Etapa 5 - Valor Justo",
                        [
                            'meses_ate_recebimento', 'ipca_mensal', 'fator_correcao_ate_recebimento',
                            'mora_ate_recebimento', 'valor_corrigido_ate_recebimento',
                            'taxa_recuperacao', 'valor_recuperavel_ate_recebimento',
                            'remuneracao_variavel_perc', 'remuneracao_variavel_valor',
                            'valor_recebimento_pos_rv', 'fator_de_desconto_vp', 'valor_justo'
                        ],
                    )

                    # ============= ETAPA 8: ADICIONAR COLUNAS INFORMATIVAS =============
                    
                    # Salvar resultado no session_state
                    st.session_state.df_final = df_final_temp
                    st.session_state.df_com_aging = df_com_aging

                    caminho_exportado, novo_arquivo = exportar_resultado_final_excel(
                        st.session_state.df_final,
                        eh_voltz=False,
                    )
                    if caminho_exportado:
                        if novo_arquivo:
                            st.success(f"💾 Resultado final exportado automaticamente em: {caminho_exportado}")
                        else:
                            st.info(f"💾 Última exportação automática: {caminho_exportado}")
                    
                    # Calcular estatísticas do DI-PRE para exibição
                    calc_valor_justo = CalculadorValorJusto()
                    stats_di_pre = calc_valor_justo.obter_estatisticas_di_pre(st.session_state.df_di_pre)
                    st.session_state.stats_di_pre_valor_justo = stats_di_pre
                    
                    progress_main.progress(1.0)
                    etapa_tempo = time.time() - etapa_inicio
                    tempo_total = time.time() - inicio_processamento
                    
                    # ========== DASHBOARD FINAL DE SUCESSO ==========
                    status_text.text("✅ Processamento ultra-otimizado concluído com sucesso!")
                    
                    # Métricas finais
                    velocidade_total = len(df_final_temp) / tempo_total if tempo_total > 0 else 0
                    
                    # Criar colunas para métricas
                    metric_col1, metric_col2, metric_col3 = st.columns(3)
                    
                    with metric_col1:
                        st.metric("📊 Registros", f"{len(df_final_temp):,}", "✅ Processados")
                    with metric_col2:
                        st.metric("⚡ Velocidade", f"{velocidade_total:,.0f}", "registros/seg")
                    with metric_col3:
                        st.metric("⏱️ Tempo Total", f"{tempo_total:.1f}s", "🎯 Ultra-rápido")
                    
                    # Dashboard de performance final
                    with log_container:
                        st.balloons()  # Animação de sucesso
                        st.success("🎉 **PROCESSAMENTO ULTRA-OTIMIZADO CONCLUÍDO!**")
                        st.info(f"📊 **Performance final:** {len(df_final_temp):,} registros processados em {tempo_total:.2f}s")
                        st.info(f"🚀 **Throughput:** {velocidade_total:,.0f} registros/segundo")
                        st.info(f"⚡ **Speedup estimado:** ~200x vs versão anterior")
                        
                        # Verificar se temos otimizações VOLTZ
                        if eh_voltz_detectado:
                            st.success("🎯 **Sistema VOLTZ:** Otimizações ultra-avançadas aplicadas com sucesso!")
                            
                            # Performance score
                            performance_score = min(100, (velocidade_total / 1000) * 100)  # Score baseado em throughput
                            st.metric("🏆 Performance Score", f"{performance_score:.0f}/100", "Ultra-Performance")
                        
                        st.success("✅ Correção monetária e valor justo calculados com sucesso!")
                        
                except Exception as e:
                    st.error(f"❌ Erro no cálculo do valor justo: {str(e)}")
                    st.warning("⚠️ Continuando com dados básicos (sem valor justo)")
                    # Salvar dados básicos mesmo com erro no valor justo
                    if df_final_temp is None:
                        df_final_temp = df_com_aging.copy() if df_com_aging is not None else pd.DataFrame()
                    st.session_state.df_final = df_final_temp
                    st.session_state.df_com_aging = df_com_aging

                    caminho_exportado, novo_arquivo = exportar_resultado_final_excel(
                        st.session_state.df_final,
                        eh_voltz=False,
                    )
                    if caminho_exportado:
                        if novo_arquivo:
                            st.success(f"💾 Dados básicos exportados automaticamente em: {caminho_exportado}")
                        else:
                            st.info(f"💾 Última exportação automática: {caminho_exportado}")

                    st.exception(e)  # Debug detalhado
                    
        except Exception as e:
            st.error(f"❌ Erro ao processar correção: {str(e)}")
            st.exception(e)  # Debug

        if 'df_final' in st.session_state and st.session_state.df_final is not None and not st.session_state.df_final.empty:
            df_resumo = st.session_state.df_final

            coluna_principal = 'valor_principal_limpo' if 'valor_principal_limpo' in df_resumo.columns else 'valor_principal'
            total_valor_principal = df_resumo[coluna_principal].sum() if coluna_principal in df_resumo.columns else 0

            if 'correcao_monetaria' in df_resumo.columns:
                total_correcao_monetaria = df_resumo['correcao_monetaria'].sum()
            else:
                total_valor_corrigido = df_resumo['valor_corrigido'].sum() if 'valor_corrigido' in df_resumo.columns else 0
                total_valor_liquido = df_resumo['valor_liquido'].sum() if 'valor_liquido' in df_resumo.columns else 0
                total_correcao_monetaria = max(total_valor_corrigido - total_valor_liquido, 0)

            coluna_valor_justo = 'valor_justo_reajustado' if 'valor_justo_reajustado' in df_resumo.columns else 'valor_justo'
            total_valor_justo = df_resumo[coluna_valor_justo].sum() if coluna_valor_justo in df_resumo.columns else 0

            st.markdown("### 📊 Comparativo Final de Valores")
            col1, col2, col3 = st.columns(3)

            with col1:
                st.metric("📊 Valor Principal", f"R$ {total_valor_principal:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

            with col2:
                st.metric("⚡ Correção Monetária", f"R$ {total_correcao_monetaria:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

            with col3:
                st.metric("💎 Valor Justo", f"R$ {total_valor_justo:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

            if coluna_valor_justo == 'valor_justo_reajustado':
                st.caption("Valor Justo exibido com base em valor_justo_reajustado (pós-RV).")

            st.write("**📋 Pré-visualização do resultado final**")
            st.dataframe(st.session_state.df_final.head(200), use_container_width=True, height=420)
            if len(st.session_state.df_final) > 200:
                st.caption(f"Mostrando 200 de {len(st.session_state.df_final):,} registros para evitar MessageSizeError.")

    # Mostrar resultados APENAS se o cálculo foi solicitado pelo usuário E temos dados calculados
    calculo_foi_solicitado = st.session_state.get('calculo_solicitado', False)
    tem_dados_calculados = 'df_final' in st.session_state and not st.session_state.df_final.empty
    
    # if calculo_foi_solicitado and tem_dados_calculados:
        
    #     # Detectar se é VOLTZ para escolher o visualizador apropriado
    #     nome_arquivo_detectado = "Distribuidora"  # Default
    #     if 'df_carregado' in st.session_state and st.session_state.df_carregado:
    #         primeiro_arquivo = list(st.session_state.df_carregado.keys())[0]
    #         nome_arquivo_detectado = primeiro_arquivo
        
    #     # Verificar se é VOLTZ
    #     from utils.calculador_voltz import CalculadorVoltz
    #     calculador_voltz = CalculadorVoltz(st.session_state.params)
    #     eh_voltz = calculador_voltz.identificar_voltz(nome_arquivo_detectado)
        
    #     # Usar o visualizador apropriado
    #     if eh_voltz:
    #         st.info("⚡ **VOLTZ detectada:** Usando visualização específica para VOLTZ")
    #         visualizador = VisualizadorVoltz()
    #         visualizador.exibir_resultados_voltz(st.session_state.df_final)
    #         visualizador.exibir_exportacao_voltz(st.session_state.df_final)
    #         visualizador.exibir_limpar_cache()
    #         visualizador.exibir_gerenciamento_checkpoints()
    #     else:
    #         st.info("🏢 **Distribuidora Geral:** Usando visualização padrão com DI-PRE")
    #         visualizador = VisualizadorDistribuidoras()
    #         visualizador.exibir_resultados_distribuidoras(st.session_state.df_final)
    #         visualizador.exibir_exportacao_distribuidoras(st.session_state.df_final)
    #         visualizador.exibir_info_processo_distribuidoras()
    #         visualizador.exibir_limpar_cache()


if __name__ == "__main__":
    show()